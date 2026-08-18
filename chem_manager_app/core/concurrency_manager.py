"""Optimistic concurrency helpers for the shared ChemicalList workbook.

The shared workbook is identified by logical row keys rather than physical row
numbers.  A sync job works on a private copy and this module compares:

* base:   the workbook as it was when the job started
* ours:   the private working copy produced by the job
* theirs: the latest locally-synchronised OneDrive copy at commit time

Only cells changed by the job are considered.  Non-overlapping edits are
merged, while overlapping edits raise ``ConcurrentEditConflict`` before save.
"""

from __future__ import annotations

from dataclasses import dataclass
import hashlib
import os
import threading
from typing import Any

from openpyxl import load_workbook

from core.db_manager import DBManager


EXCLUDED_MAIN_SHEETS = {
    "db", "guide", "가이드", "가이드(guide)", "index", "old chemical list"
}

# All workbook writers in this process share this lock.  OneDrive concurrency
# is still handled optimistically; this lock only prevents our own workers from
# racing each other.
WORKBOOK_OPERATION_LOCK = threading.Lock()


class ConcurrentEditConflict(RuntimeError):
    """Raised when both the program and a coauthor changed the same cell."""

    def __init__(self, conflicts: list[str]):
        self.conflicts = conflicts
        preview = "; ".join(conflicts[:5])
        if len(conflicts) > 5:
            preview += f" 외 {len(conflicts) - 5}건"
        super().__init__(f"공동편집 충돌이 감지되었습니다: {preview}")


@dataclass(frozen=True)
class FileFingerprint:
    mtime_ns: int
    size: int
    sha256: str


@dataclass
class SheetSnapshot:
    name: str
    key_kind: str
    headers: list[str]
    rows: dict[str, dict[str, Any]]


@dataclass(frozen=True)
class CellPatch:
    sheet: str
    key_kind: str
    row_key: str
    column: str
    base_value: Any
    new_value: Any
    new_row: bool = False


def file_fingerprint(path: str) -> FileFingerprint:
    stat = os.stat(path)
    digest = hashlib.sha256()
    with open(path, "rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return FileFingerprint(stat.st_mtime_ns, stat.st_size, digest.hexdigest())


def values_equal(left: Any, right: Any) -> bool:
    if left is None:
        left = ""
    if right is None:
        right = ""
    if isinstance(left, str):
        left = left.strip()
    if isinstance(right, str):
        right = right.strip()
    return left == right


def _normalise_catalog(value: Any) -> str:
    text = str(value or "").strip()
    if text.endswith(".0"):
        text = text[:-2]
    return text


def logical_key(key_kind: str, row: dict[str, Any]) -> str:
    if key_kind == "order":
        return str(row.get("Order No.", "") or "").strip()
    manufacturer = DBManager.normalize_manufacturer(row.get("Manufacturer", ""))
    catalog = _normalise_catalog(row.get("Catalog No.", ""))
    if not manufacturer or not catalog:
        return ""
    return f"{manufacturer.lower()}|{catalog}"


def legacy_order_key(row: dict[str, Any], occurrence: int = 1, row_number: int = 0) -> str:
    """Stable fallback key for legacy ChemicalList rows without Order No."""
    manufacturer = DBManager.normalize_manufacturer(row.get("Manufacturer", "")).casefold()
    catalog = _normalise_catalog(row.get("Catalog No.", "")).casefold()
    lot = str(row.get("Lot No.", "") or "").strip().casefold()
    original_value = row.get("Original Product Name", "") or row.get("Product Name", "") or ""
    original_name = "" if str(original_value).startswith("=") else str(original_value).strip().casefold()
    identity = "|".join((manufacturer, catalog, lot, original_name))
    if any((manufacturer, catalog, lot, original_name)):
        return f"__legacy__|{identity}|#{occurrence}"
    return f"__legacy_row__|{row_number}"


def legacy_row_has_data(row: dict[str, Any]) -> bool:
    """Ignore reserved rows whose only content is a program formula."""
    for value in row.values():
        if value in (None, ""):
            continue
        if isinstance(value, str) and value.startswith("="):
            continue
        return True
    return False


def _select_sheets(workbook) -> list[tuple[Any, str]]:
    selected: list[tuple[Any, str]] = []
    for worksheet in workbook.worksheets:
        lowered = worksheet.title.strip().lower()
        if lowered == "db":
            selected.append((worksheet, "db"))
        elif lowered not in EXCLUDED_MAIN_SHEETS and not any(kind == "order" for _, kind in selected):
            selected.append((worksheet, "order"))
    return selected


def snapshot_workbook(path: str) -> dict[str, SheetSnapshot]:
    workbook = load_workbook(path, read_only=True, data_only=False, keep_links=False)
    snapshots: dict[str, SheetSnapshot] = {}
    try:
        for worksheet, key_kind in _select_sheets(workbook):
            header_values = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
            headers = [str(value).strip() if value is not None else "" for value in header_values]
            rows: dict[str, dict[str, Any]] = {}
            legacy_occurrences: dict[str, int] = {}
            for row_number, values in enumerate(
                worksheet.iter_rows(min_row=2, values_only=True), start=2
            ):
                row = {
                    header: values[index] if index < len(values) else None
                    for index, header in enumerate(headers)
                    if header
                }
                key = logical_key(key_kind, row)
                if key_kind == "order" and not key and legacy_row_has_data(row):
                    identity = legacy_order_key(row, occurrence=1, row_number=row_number).rsplit("|#", 1)[0]
                    legacy_occurrences[identity] = legacy_occurrences.get(identity, 0) + 1
                    key = legacy_order_key(
                        row, legacy_occurrences[identity], row_number=row_number
                    )
                if key and key not in rows:
                    rows[key] = row
            snapshots[key_kind] = SheetSnapshot(
                name=worksheet.title,
                key_kind=key_kind,
                headers=[header for header in headers if header],
                rows=rows,
            )
    finally:
        workbook.close()
    return snapshots


def build_three_way_plan(
    base: dict[str, SheetSnapshot],
    ours: dict[str, SheetSnapshot],
    theirs: dict[str, SheetSnapshot],
) -> tuple[list[CellPatch], list[str]]:
    patches: list[CellPatch] = []
    conflicts: list[str] = []

    for key_kind, our_sheet in ours.items():
        base_sheet = base.get(key_kind)
        their_sheet = theirs.get(key_kind)
        if their_sheet is None:
            if base_sheet is not None:
                conflicts.append(f"{our_sheet.name}: 공동편집자가 시트를 삭제함")
                continue
            their_sheet = SheetSnapshot(
                name=our_sheet.name,
                key_kind=key_kind,
                headers=[],
                rows={},
            )
        if base_sheet and base_sheet.headers != their_sheet.headers:
            conflicts.append(f"{their_sheet.name}: 공동편집자가 헤더 구조를 변경함")
            continue

        base_rows = base_sheet.rows if base_sheet else {}
        their_rows = their_sheet.rows
        for row_key, our_row in our_sheet.rows.items():
            base_row = base_rows.get(row_key)
            their_row = their_rows.get(row_key)

            if base_row is None:
                if their_row is None:
                    for column, new_value in our_row.items():
                        if column:
                            patches.append(CellPatch(
                                their_sheet.name, key_kind, row_key, column,
                                None, new_value, new_row=True,
                            ))
                else:
                    for column, new_value in our_row.items():
                        if not column or values_equal(new_value, ""):
                            continue
                        their_value = their_row.get(column)
                        if values_equal(their_value, ""):
                            patches.append(CellPatch(
                                their_sheet.name, key_kind, row_key, column,
                                None, new_value, new_row=False,
                            ))
                        elif not values_equal(their_value, new_value):
                            conflicts.append(f"{their_sheet.name}[{row_key}].{column}")
                continue

            if their_row is None:
                conflicts.append(f"{their_sheet.name}[{row_key}]: 공동편집자가 행을 삭제함")
                continue

            columns = set(base_row) | set(our_row)
            for column in columns:
                base_value = base_row.get(column)
                our_value = our_row.get(column)
                if values_equal(base_value, our_value):
                    continue
                their_value = their_row.get(column)
                if values_equal(their_value, base_value):
                    patches.append(CellPatch(
                        their_sheet.name, key_kind, row_key, column,
                        base_value, our_value, new_row=False,
                    ))
                elif not values_equal(their_value, our_value):
                    conflicts.append(f"{their_sheet.name}[{row_key}].{column}")

    return patches, conflicts


def format_validation_max_row(last_data_row: int) -> int:
    """Return the managed boundary: 3,000 rows, then 500-row increments."""
    last_row = max(1, int(last_data_row or 1))
    if last_row <= 2500:
        return 3000
    required = last_row + 1000
    return ((required + 499) // 500) * 500


def workbook_range_needs_extension(path: str, required_row: int) -> bool:
    """Check whether program-managed formatting reaches ``required_row``.

    Excel may store list validation as an x14 extension which openpyxl cannot
    expose (it reports an empty validation collection).  Conditional formatting
    is written in the same operation and to the same managed row boundary, so it
    is the stable sentinel.  Requiring both caused every no-change sync to create
    another backup even though the range was already current.
    """
    workbook = load_workbook(path, read_only=False, data_only=False, keep_links=False)
    try:
        selected = _select_sheets(workbook)
        main_sheet = next((sheet for sheet, kind in selected if kind == "order"), None)
        if main_sheet is None:
            return True

        conditional_max = 0
        rules = getattr(main_sheet.conditional_formatting, "_cf_rules", {})
        for conditional_range in rules:
            for cell_range in conditional_range.sqref.ranges:
                conditional_max = max(conditional_max, cell_range.max_row)

        return conditional_max < required_row
    finally:
        workbook.close()


def workbook_main_sheet_needs_activation(path: str) -> bool:
    """Return True when the workbook would open on a non-ChemicalList sheet."""
    workbook = load_workbook(path, read_only=True, data_only=False, keep_links=False)
    try:
        active = workbook.active
        return active is None or active.title.strip().casefold() != "chemicallist"
    finally:
        workbook.close()
