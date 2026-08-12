import os
import pandas as pd
from datetime import datetime
import re
import openpyxl

class DBManager:
    COLUMN_MAP = {
        "제조사": "Manufacturer",
        "제품번호": "Catalog No.",
        "시약명": "Product Name",
        "CAS Number": "CAS No.",
        "CAS 번호": "CAS No.",
        "보관온도": "Storage Temp.",
        "신호어": "Signal Word",
        "주요위험": "Key Hazards",
        "상세 위험분류": "Detailed Hazard Classification",
        "민감성": "Sensitivity",
        "상세정보_링크": "Detail_Link",
        "갱신일": "Revision Date"
    }

    def __init__(self, db_path="ChemicalList.xlsx", sheet_name="DB"):
        self.db_path = db_path
        self.sheet_name = sheet_name
        self.columns = [
            "Key", "Manufacturer", "Catalog No.", "Product Name", "CAS No.", 
            "Storage Temp.", "Sensitivity", "Signal Word", "Key Hazards", "Detailed Hazard Classification", "Detail_Link", "SDS_Link", 
            "SDS_Local_Path", "Revision Date"
        ]
        self._init_db()

    def _init_db(self):
        """DB 파일/시트가 없으면 초기 템플릿을 생성합니다."""
        if not os.path.exists(self.db_path):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = self.sheet_name
            ws.append(self.columns)
            wb.save(self.db_path)
            wb.close()
        else:
            wb = openpyxl.load_workbook(self.db_path)
            if self.sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(title=self.sheet_name)
                ws.append(self.columns)
                wb.save(self.db_path)
            wb.close()

    def load_db(self):
        """DB 시트를 pandas DataFrame으로 로드하고 영문 표준 헤더로 정규화합니다."""
        self._init_db()
        df = pd.read_excel(self.db_path, sheet_name=self.sheet_name)
        df = df.rename(columns=self.COLUMN_MAP)
        df = df.loc[:, ~df.columns.duplicated()]
        return df

    def save_db(self, df):
        """DataFrame을 1) 제조사 2) 시약명 오름차순으로 정렬하여 DB 시트에 저장합니다."""
        df = df.rename(columns=self.COLUMN_MAP)
        df = df.loc[:, ~df.columns.duplicated()]
        
        for col in self.columns:
            if col not in df.columns:
                df[col] = ""

        # Filter strictly to self.columns in order, dropping extra/Korean columns
        df = df[self.columns]

        m_col = "Manufacturer"
        p_col = "Product Name"
        df = df.sort_values(by=[m_col, p_col], ascending=[True, True])
        
        self._init_db()
        wb = openpyxl.load_workbook(self.db_path)
        
        if self.sheet_name in wb.sheetnames:
            ws = wb[self.sheet_name]
            if ws.max_row >= 1:
                ws.delete_rows(1, ws.max_row + 1)
        else:
            ws = wb.create_sheet(title=self.sheet_name)
            
        headers = self.columns
        ws.append(headers)
        
        for row_idx, row in enumerate(df.itertuples(index=False), start=2):
            cleaned_row = [None if pd.isna(val) else str(val).strip() for val in row]
            # Col 1 (Key): Excel formula
            cleaned_row[0] = f'=B{row_idx}&"|"&C{row_idx}'
            # Col 14 (Revision Date): Date only
            rev_val = cleaned_row[13] or ""
            if rev_val and " " in str(rev_val):
                cleaned_row[13] = str(rev_val).split(" ")[0]
            ws.append(cleaned_row)
            
            # Format Catalog No. (Col 3) as text format
            ws.cell(row=row_idx, column=3).number_format = "@"
            
        # Apply explicit formatting to DB Sheet
        from openpyxl.styles import PatternFill, Font, Alignment
        
        header_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        header_font = Font(bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")

        # 1. Header Row: Bold, RGB(226, 239, 218), Center Alignment
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        # 2. Data Rows: Cols 1..6 Center, Cols 7+ Left Alignment
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.column <= 6:
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align

        header_row = next(ws.iter_rows(min_row=1, max_row=1))
        for col_idx, column in enumerate(ws.columns, 1):
            col_name = str(header_row[col_idx-1].value)
            if col_name in ['Product Name', '시약명']:
                ws.column_dimensions[column[0].column_letter].width = 45
            elif col_name in ['Detailed Hazard Classification', '상세 위험분류']:
                ws.column_dimensions[column[0].column_letter].width = 50
            elif col_name in ['Key Hazards', 'Signal Word', '주요위험', '신호어']:
                ws.column_dimensions[column[0].column_letter].width = 25
        
        import time
        max_retries = 5
        for attempt in range(max_retries):
            try:
                wb.save(self.db_path)
                wb.close()
                return
            except (PermissionError, OSError) as pe:
                if attempt < max_retries - 1:
                    time.sleep(0.5)
                else:
                    wb.close()
                    raise Exception(f"ChemicalList.xlsx 파일이 다른 프로그램(예: 엑셀)에서 열려 있어 저장할 수 없습니다.\n엑셀에서 ChemicalList.xlsx 파일을 닫은 후 다시 [DB 수동 업데이트]를 시도해 주세요.\n({pe})")

    def get_product(self, manufacturer, product_number):
        """특정 제조사와 제품번호의 데이터를 DB에서 검색합니다."""
        df = self.load_db()
        m_col = "Manufacturer"
        c_col = "Catalog No."
        mask = (
            (df[m_col].astype(str).str.lower().str.strip() == str(manufacturer).lower().strip()) & 
            (df[c_col].astype(str).str.strip() == str(product_number).strip())
        )
        result = df[mask]
        
        if not result.empty:
            return result.iloc[0].to_dict()
        return None

    def add_product(self, data_dict):
        """수집된 제품 데이터를 DB에 추가/업데이트합니다."""
        self.add_products_batch([data_dict])

    def add_products_batch(self, data_dict_list):
        """수집된 여러 제품 데이터 리스트를 DB에 한 번에 추가/업데이트합니다."""
        if not data_dict_list:
            return
            
        df = self.load_db()
        
        normalized_list = []
        for data_dict in data_dict_list:
            norm_dict = {}
            for k, v in data_dict.items():
                norm_k = self.COLUMN_MAP.get(k, k)
                norm_dict[norm_k] = v
                
            norm_dict["Revision Date"] = datetime.now().strftime("%Y-%m-%d")
            if "Manufacturer" in norm_dict:
                norm_dict["Manufacturer"] = DBManager.normalize_manufacturer(norm_dict["Manufacturer"])
            normalized_list.append(norm_dict)

        new_rows = pd.DataFrame(normalized_list)
        
        for norm_dict in normalized_list:
            m_val = str(norm_dict.get("Manufacturer", "")).strip().lower()
            c_val = str(norm_dict.get("Catalog No.", "")).strip()
            
            if "Manufacturer" in df.columns and "Catalog No." in df.columns:
                mask = (
                    (df["Manufacturer"].astype(str).str.lower().str.strip() == m_val) & 
                    (df["Catalog No."].astype(str).str.strip() == c_val)
                )
                df = df[~mask]
                
        df = pd.concat([df, new_rows], ignore_index=True)
        self.save_db(df)

    @staticmethod
    def normalize_manufacturer(m_name):
        if not m_name or pd.isna(m_name):
            return ""
        s = str(m_name).strip()
        s_lower = s.lower()
        
        # Aldrich family: aldrich, sigma-aldrich, sigma, merck, sial, millipore, etc.
        aldrich_kw = ["aldrich", "sigma", "merck", "sial", "millipore", "머크", "알드리치", "시그마"]
        if any(kw in s_lower for kw in aldrich_kw):
            return "Aldrich"
            
        # Thermo family
        thermo_kw = ["thermo", "alfa", "fisher", "invitrogen", "acros", "써모", "피셔"]
        if any(kw in s_lower for kw in thermo_kw):
            return "ThermoFisher"
            
        # TCI family
        tci_kw = ["tci", "tokyo kasei", "티씨아이", "도쿄카세이"]
        if any(kw in s_lower for kw in tci_kw):
            return "TCI"
            
        # Abcam
        if "abcam" in s_lower or "앱캠" in s_lower:
            return "Abcam"
            
        return s.title() if len(s) > 1 else s.upper()

    @staticmethod
    def clean_filename(filename):
        """파일명으로 사용할 수 없는 특수문자를 제거합니다."""
        if not filename or not isinstance(filename, str) or filename == "제조사 홈페이지에서 검색 실패":
            return "unknown"
        cleaned = re.sub(r'[\\/*?:"<>|\r\n]', "", str(filename))
        cleaned = cleaned.strip()
        if not cleaned:
            return "unknown"
        return cleaned

    @staticmethod
    def format_sds_filename(product_name, manufacturer, catalog_no):
        """Formats SDS filename as: Product Name (Manufacturer, Catalog No).pdf"""
        p_name = DBManager.clean_filename(product_name)
        mfr = DBManager.clean_filename(manufacturer)
        cat = DBManager.clean_filename(catalog_no)
        
        if p_name == "unknown" or not p_name:
            p_name = "Product"
        if mfr == "unknown" or not mfr:
            mfr = "Manufacturer"
        if cat == "unknown" or not cat:
            cat = "CatalogNo"
            
        return f"{p_name} ({mfr}, {cat})"

    @staticmethod
    def normalize_temperature(temp_text):
        if not temp_text or pd.isna(temp_text):
            return "-"
            
        t_str = str(temp_text).strip()
        t_lower = t_str.lower()

        if t_lower in ["none", "nan", "n/a", "정보 없음", "검색 실패", "search failed", "", "unknown", "-"]:
            return "-"

        # If already one of the strict valid codes
        if t_str in ["RT", "R", "F", "DF"]:
            return t_str

        # 1. Keywords check for Cold / Frozen / Deep Freezer storage FIRST (to avoid false RT matches inside 'refrigerated')
        if any(kw in t_lower for kw in ["deep freezer", "ultracold", "deep-freezer", "초저온"]):
            return "DF"
        if any(kw in t_lower for kw in ["frozen", "freezer", "freeze", "냉동"]):
            return "F"
        if any(kw in t_lower for kw in ["refrigerat", "refrigerator", "cold", "chilled", "cool", "냉장"]):
            return "R"

        # 2. Keywords check for Room Temperature (using word boundaries for 'rt')
        if any(kw in t_lower for kw in ["room temp", "ambient", "실온", "상온"]) or re.search(r'\brt\b', t_lower):
            return "RT"

        # 3. Standardize range hyphens/tildes between digits so "2-30" -> "2 to 30", "2-20" -> "2 to 20"
        normalized = re.sub(r'(\d+)\s*[\-~～to]+\s*(\d+)', r'\1 to \2', t_str)

        # 4. Parse numbers for temperature range
        numbers = [float(n) for n in re.findall(r'[-+]?\d+(?:\.\d+)?', normalized)]
        
        if numbers:
            max_temp = max(numbers)
            if max_temp >= 21:
                return "RT"
            elif max_temp >= 0:
                return "R"
            elif max_temp >= -30:
                return "F"
            else:
                return "DF"

        return "-"

    @staticmethod
    def extract_sensitivity(text):
        if not text or pd.isna(text): return "-"
        t = str(text).lower()
        if t in ["none", "nan", "n/a", "정보 없음", "검색 실패", "search failed", "", "-"]:
            return "-"
            
        sensitivities = []
        
        if "moisture" in t or "hygroscopic" in t or "흡습" in t or "수분" in t:
            sensitivities.append("Moisture")
        if "light" in t or "광" in t or "차광" in t:
            sensitivities.append("Light")
        if "air" in t or "oxygen" in t or "공기" in t or "산소" in t:
            sensitivities.append("Air")
        if "heat" in t or "열" in t:
            sensitivities.append("Heat")
        if "argon" in t or "nitrogen" in t or "불활성" in t:
            sensitivities.append("Ar/N2 Required")
            
        if not sensitivities:
            return "-"
        return ", ".join(sensitivities)
