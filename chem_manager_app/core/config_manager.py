import json
import os

import sys
import copy

def get_app_root():
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

CONFIG_FILE = os.path.join(get_app_root(), "config.json")

DEFAULT_CONFIG = {
    "source_file": "",
    "target_file": "",
    "source_sheet": "",
    "header_row": 1,
    "source_headers": [
        "번호", "날짜", "주문자", "회사", "품목명", 
        "수령확인", "CAS 번호", "품번", "용량", "수량", "보관온도", "Lot No."
    ],
    "target_headers": [
        "Order No.", "Order Date", "Ordered By", "Product Name", 
        "Manufacturer", "Package Size", "CAS No.", "Catalog No.", 
        "Room", "Storage Temp.", "Cabinet", "Quantity", "Used", "Status", "Remarks",
        "Lot No.", "Expiration Date", "COA Link", "COA Local Path"
    ],
    "mapping": {
        "Order No.": "번호",
        "Order Date": "날짜",
        "Ordered By": "주문자",
        "Product Name": "품목명",
        "Manufacturer": "회사",
        "Package Size": "용량",
        "CAS No.": "CAS 번호",
        "Catalog No.": "품번",
        "Room": "",
        "Storage Temp.": "보관온도",
        "Cabinet": "",
        "Quantity": "수량",
        "Used": "",
        "Status": "",
        "Remarks": "",
        "Lot No.": "Lot No."
    },
    "colors": {
        "done_bg": "#e6f7e6",
        "done_font": "#000000",
        "warn_cas_bg": "#fffbe6",
        "warn_cas_font": "#000000",
        "warn_num_bg": "#fff2e6",
        "warn_num_font": "#000000",
        "warn_both_bg": "#ffe6e6",
        "warn_both_font": "#000000",
        "warn_loc_bg": "#e6f2ff",
        "warn_loc_font": "#000000",
        "normal_bg": "#ffffff",
        "normal_font": "#000000",
        "header_bg": "#d9d9d9",
        "header_font": "#000000",
        "search_failed_bg": "#ffff00",
        "search_failed_font": "#ff0000",
        "manual_input_bg": "#ffff00",
        "manual_input_font": "#ff0000",
        "status_x_bg": "#ffe6e6",
        "status_x_font": "#ff0000"
    },
    "sync_interval_minutes": 0,
    "watch_enabled": False,
    "run_on_startup": False
}

def normalize_local_path(value):
    """Return a normalized local path, including file:/// values."""
    import urllib.parse

    text = str(value or "").strip()
    if text.startswith("file:///"):
        text = urllib.parse.unquote(text[8:])
    elif text.startswith("file://"):
        text = urllib.parse.unquote(text[7:])
    return os.path.abspath(text) if text else ""


def resolve_target_file(config):
    """Resolve the configured ChemicalList workbook with legacy fallback."""
    configured = normalize_local_path((config or {}).get("target_file", ""))
    if configured:
        return configured
    source = normalize_local_path((config or {}).get("source_file", ""))
    if source:
        return os.path.join(os.path.dirname(source), "ChemicalList.xlsx")
    return os.path.abspath("ChemicalList.xlsx")


def validate_chemical_list_file(path):
    """Validate an existing target workbook without opening it in Excel."""
    normalized = normalize_local_path(path)
    if not normalized or not os.path.isfile(normalized):
        return False, "파일을 찾을 수 없습니다."
    if not normalized.lower().endswith(".xlsx"):
        return False, "ChemicalList 파일은 .xlsx 형식이어야 합니다."
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(normalized, read_only=True, data_only=False, keep_links=False)
        try:
            names = set(workbook.sheetnames)
        finally:
            workbook.close()
    except Exception as error:
        return False, f"Excel 파일을 읽을 수 없습니다: {error}"
    missing = [name for name in ("ChemicalList", "DB") if name not in names]
    if missing:
        return False, f"필수 시트가 없습니다: {', '.join(missing)}"
    return True, ""

def load_config():
    if not os.path.exists(CONFIG_FILE):
        save_config(DEFAULT_CONFIG)
        return copy.deepcopy(DEFAULT_CONFIG)
    
    try:
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
            # Merge with defaults for missing keys
            for key, val in DEFAULT_CONFIG.items():
                if key not in data:
                    data[key] = val
                elif isinstance(val, dict):
                    # For nested dictionaries like mapping and colors
                    for sub_key, sub_val in val.items():
                        if sub_key not in data[key]:
                            data[key][sub_key] = sub_val
            if "Lot No." not in data["source_headers"]:
                data["source_headers"].append("Lot No.")
            for header in ("Lot No.", "Expiration Date", "COA Link", "COA Local Path"):
                if header not in data["target_headers"]:
                    data["target_headers"].append(header)
            data["mapping"].setdefault("Lot No.", "Lot No.")
            return data
    except Exception as e:
        raise OSError(f"설정 파일을 읽을 수 없습니다: {CONFIG_FILE} ({e})") from e

def save_config(config_data):
    try:
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(config_data, f, ensure_ascii=False, indent=4)
        return True
    except Exception as e:
        raise OSError(f"설정 파일을 저장할 수 없습니다: {CONFIG_FILE} ({e})") from e
