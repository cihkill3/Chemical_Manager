import os
import time
import shutil
import datetime
import re
import tempfile
import win32com.client
from utils.color_utils import hex_to_bgr
from utils.excel_utils import get_col_letter
from core.concurrency_manager import (
    ConcurrentEditConflict,
    build_three_way_plan,
    file_fingerprint,
    format_validation_max_row,
    legacy_order_key,
    legacy_row_has_data,
    snapshot_workbook,
    workbook_range_needs_extension,
    workbook_main_sheet_needs_activation,
)
from core.excel_manager import (
    build_header_map,
    check_and_wait_lock,
    is_file_locked,
    meaningful_data_last_row,
    reserve_missing_headers,
    should_manage_product_name,
    write_reserved_headers,
)
from core.coa_manager import COAManager, COA_HEADERS, is_supported_vendor, lots_equal, normalize_lot, valid_cached_document
from core.backup_manager import create_backup, purge_expired_backups
from core.config_manager import normalize_local_path, resolve_target_file, validate_chemical_list_file

class SyncEngine:
    TEXT_FIELDS = {"Manufacturer", "Catalog No."}

    @staticmethod
    def as_excel_text(value):
        """Return identifier-like Excel values without COM's artificial .0."""
        if value is None:
            return ""
        if isinstance(value, bool):
            return str(value)
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    @staticmethod
    def is_received(value):
        """Accept O/o, the Korean circle, or a valid supported receipt date."""
        if isinstance(value, (datetime.datetime, datetime.date)):
            return True
        text = SyncEngine.as_excel_text(value)
        if text.upper() == "O" or text == "ㅇ":
            return True
        for pattern, date_format in (
            (r"\d{4}-\d{2}-\d{2}", "%Y-%m-%d"),
            (r"\d{2}-\d{2}-\d{2}", "%y-%m-%d"),
            (r"\d{6}", "%y%m%d"),
        ):
            if re.fullmatch(pattern, text):
                try:
                    datetime.datetime.strptime(text, date_format)
                    return True
                except ValueError:
                    return False
        return False

    @staticmethod
    def get_col_idx(names, col_map):
        for n in names:
            if n in col_map and col_map[n] > 0:
                return col_map[n]
        return 0

    def __init__(self, config_data, callback_progress=None, check_stop_fn=None):
        self.config = config_data
        self.callback = callback_progress
        self.check_stop_fn = check_stop_fn
        self.excel = None
        self.source_wb = None
        self.target_wb = None
        self.is_stopped_flag = False
        self._working_copy_path = None
        self._base_snapshot = None
        self._base_fingerprint = None
        self._restart_count = 0
        self._coa_commit_metadata = []

    def is_stopped(self):
        if self.is_stopped_flag:
            return True
        if self.check_stop_fn:
            return self.check_stop_fn()
        return False

    def log(self, message):
        if self.callback:
            self.callback(message)
        else:
            print(message)

    def safe_set_val(self, cell_or_range, val, retries=5):
        for attempt in range(retries):
            try:
                cell_or_range.Value = val
                return
            except Exception as e:
                if attempt == retries - 1:
                    raise e
                time.sleep(0.1)

    def _auto_sync_limits(self):
        interval = max(0, int(self.config.get("sync_interval_minutes", 0) or 0))
        is_auto_sync = interval > 0
        return is_auto_sync, (max(1, interval - 1) if is_auto_sync else None)

    def _discard_working_copy(self):
        if self._working_copy_path and os.path.exists(self._working_copy_path):
            try:
                os.remove(self._working_copy_path)
            except Exception:
                pass
        self._working_copy_path = None

    def _make_working_copy(self, target_path):
        """Copy the shared workbook outside OneDrive before any long-running work."""
        self._discard_working_copy()
        fd, work_path = tempfile.mkstemp(prefix="ChemicalList_work_", suffix=".xlsx")
        os.close(fd)
        os.remove(work_path)
        before = file_fingerprint(target_path)
        shutil.copy2(target_path, work_path)
        after = file_fingerprint(target_path)
        if before != after:
            os.remove(work_path)
            raise ConcurrentEditConflict(["기준본 복사 중 OneDrive 파일이 변경됨"])
        self._working_copy_path = work_path
        self._base_fingerprint = after
        self._base_snapshot = snapshot_workbook(work_path)
        return work_path

    @staticmethod
    def _sheet_key_kind(sheet_name):
        return "db" if str(sheet_name).strip().lower() == "db" else "order"

    def _apply_cell_patches(self, workbook, patches):
        grouped = {}
        for patch in patches:
            grouped.setdefault((patch.sheet, patch.key_kind), []).append(patch)

        for (sheet_name, key_kind), sheet_patches in grouped.items():
            worksheet = None
            for candidate in workbook.Worksheets:
                if candidate.Name == sheet_name:
                    worksheet = candidate
                    break
            if worksheet is None:
                worksheet = workbook.Worksheets.Add(After=workbook.Worksheets(workbook.Worksheets.Count))
                worksheet.Name = sheet_name
            last_col = worksheet.Cells(1, worksheet.Columns.Count).End(-4159).Column
            header_map = {}
            for column in range(1, max(last_col + 1, 30)):
                value = worksheet.Cells(1, column).Value
                if value:
                    header_map[str(value).strip()] = column

            from core.db_manager import DBManager
            required_headers = {patch.column for patch in sheet_patches if patch.column}
            preferred_headers = (
                list(DBManager.COLUMNS) if key_kind == "db"
                else list(self.config.get("target_headers", [])) + ["Original Product Name"]
            )
            ordered_headers = [header for header in preferred_headers if header in required_headers]
            ordered_headers.extend(sorted(required_headers - set(ordered_headers)))
            for header in ordered_headers:
                if header not in header_map:
                    last_col += 1
                    worksheet.Cells(1, last_col).Value = header
                    header_map[header] = last_col

            key_rows = {}
            last_row = meaningful_data_last_row(worksheet, header_map, key_kind)
            order_col = header_map.get("Order No.", 0)
            manufacturer_col = header_map.get("Manufacturer", 0)
            catalog_col = header_map.get("Catalog No.", 0)
            legacy_occurrences = {}
            for row in range(2, last_row + 1):
                if key_kind == "order" and order_col:
                    key = str(worksheet.Cells(row, order_col).Value or "").strip()
                    if not key:
                        row_values = {}
                        for header, column in header_map.items():
                            cell = worksheet.Cells(row, column)
                            formula = str(getattr(cell, "Formula", "") or "")
                            row_values[header] = formula if formula.startswith("=") else cell.Value
                        if legacy_row_has_data(row_values):
                            identity = legacy_order_key(
                                row_values, occurrence=1, row_number=row
                            ).rsplit("|#", 1)[0]
                            legacy_occurrences[identity] = legacy_occurrences.get(identity, 0) + 1
                            key = legacy_order_key(
                                row_values, legacy_occurrences[identity], row_number=row
                            )
                elif key_kind == "db" and manufacturer_col and catalog_col:
                    manufacturer = DBManager.normalize_manufacturer(self.as_excel_text(worksheet.Cells(row, manufacturer_col).Value))
                    catalog = self.as_excel_text(worksheet.Cells(row, catalog_col).Value)
                    key = f"{manufacturer.lower()}|{catalog}" if manufacturer and catalog else ""
                else:
                    key = ""
                if key and key not in key_rows:
                    key_rows[key] = row

            patches_by_key = {}
            for patch in sheet_patches:
                patches_by_key.setdefault(patch.row_key, []).append(patch)

            for row_key, row_patches in patches_by_key.items():
                if self.is_stopped():
                    raise Exception("사용자에 의해 저장 작업이 중단되었습니다.")
                row = key_rows.get(row_key)
                if row is None:
                    last_row += 1
                    row = last_row
                    key_rows[row_key] = row
                for patch in row_patches:
                    column = header_map.get(patch.column)
                    if column:
                        value = patch.new_value
                        if patch.column in self.TEXT_FIELDS or patch.column == "Revision Date":
                            value = self.as_excel_text(value)
                            worksheet.Cells(row, column).NumberFormat = "@"
                        if patch.column == "Revision Date":
                            from core.db_manager import DBManager
                            value = DBManager.normalize_revision_date(value)
                        self.safe_set_val(worksheet.Cells(row, column), value)

    def _apply_db_lookup_formulas(self, worksheet):
        """Rebuild program-owned DB formulas against each row's actual location."""
        header_map, _ = build_header_map(worksheet)
        manufacturer_col = header_map.get("Manufacturer", 0)
        catalog_col = header_map.get("Catalog No.", 0)
        order_col = header_map.get("Order No.", 0)
        if not manufacturer_col or not catalog_col:
            return

        db_ws = next(
            (sheet for sheet in worksheet.Parent.Worksheets
             if str(sheet.Name).strip().lower() == "db"),
            None,
        )
        if db_ws is None:
            return
        db_headers, _ = build_header_map(db_ws)
        last_row = meaningful_data_last_row(worksheet, header_map, "order")
        mfr_letter = get_col_letter(manufacturer_col)
        catalog_letter = get_col_letter(catalog_col)
        db_key_col = db_headers.get("Key", 0)
        if not db_key_col:
            return
        db_key_letter = get_col_letter(db_key_col)

        def lookup_expression(field, row):
            target_letter = get_col_letter(db_headers[field])
            return (
                f'INDEX(DB!${target_letter}:${target_letter},MATCH('
                f'{mfr_letter}{row}&"|"&{catalog_letter}{row},'
                f'DB!${db_key_letter}:${db_key_letter},0))'
            )

        lookup_fields = (
            "Sensitivity", "Signal Word", "Key Hazards",
            "Detailed Hazard Classification", "Detail_Link", "SDS_Link",
            "SDS_Local_Path",
        )
        for row in range(2, last_row + 1):
            manufacturer = self.as_excel_text(worksheet.Cells(row, manufacturer_col).Value)
            catalog = self.as_excel_text(worksheet.Cells(row, catalog_col).Value)
            if not manufacturer or not catalog:
                continue

            for field in lookup_fields:
                target_col = header_map.get(field, 0)
                db_col = db_headers.get(field, 0)
                if target_col and db_col:
                    lookup = lookup_expression(field, row)
                    formula = (
                        f'=IF(OR({mfr_letter}{row}="", {catalog_letter}{row}=""), "", '
                        f'IFERROR(IF({lookup}="","-",{lookup}), "-"))'
                    )
                    self.safe_set_val(worksheet.Cells(row, target_col), formula)

            original_col = header_map.get("Original Product Name", 0)
            product_col = header_map.get("Product Name", 0)
            db_product_col = db_headers.get("Product Name", 0)
            product_formula = (
                str(getattr(worksheet.Cells(row, product_col), "Formula", "") or "")
                if product_col else ""
            )
            order_number = worksheet.Cells(row, order_col).Value if order_col else ""
            if (
                original_col and product_col and db_product_col
                and should_manage_product_name(order_number, product_formula)
            ):
                original = worksheet.Cells(row, original_col).Value
                if not original or not str(original).strip():
                    current = worksheet.Cells(row, product_col).Value
                    if current and not str(current).startswith("="):
                        worksheet.Cells(row, original_col).Value = current
                original_letter = get_col_letter(original_col)
                lookup = lookup_expression("Product Name", row)
                formula = (
                    f'=IF(OR({mfr_letter}{row}="", {catalog_letter}{row}=""), {original_letter}{row}, '
                    f'IFERROR(IF({lookup}="", {original_letter}{row}, {lookup}), {original_letter}{row}))'
                )
                self.safe_set_val(worksheet.Cells(row, product_col), formula)

            cas_col = header_map.get("CAS No.", 0)
            db_cas_col = db_headers.get("CAS No.", 0)
            if cas_col and db_cas_col:
                cell = worksheet.Cells(row, cas_col)
                current_formula = str(getattr(cell, "Formula", "") or "")
                current_value = self.as_excel_text(cell.Value)
                replaceable = (
                    current_formula.startswith("=") or not current_value or
                    current_value in {"None", "nan", "N/A", "Search Failed", "Manual Input Required", "Manual Entry Required"}
                )
                if replaceable:
                    lookup = lookup_expression("CAS No.", row)
                    formula = (
                        f'=IF(OR({mfr_letter}{row}="", {catalog_letter}{row}=""), "", '
                        f'IFERROR({lookup}, ""))'
                    )
                    self.safe_set_val(cell, formula)

            storage_col = header_map.get("Storage Temp.", 0)
            db_storage_col = db_headers.get("Storage Temp.", 0)
            if storage_col and db_storage_col:
                lookup = lookup_expression("Storage Temp.", row)
                formula = (
                    f'=IFERROR(IF(OR({lookup}="", {lookup}="-"), "", {lookup}), "")'
                )
                self.safe_set_val(worksheet.Cells(row, storage_col), formula)

    def _apply_coa_commit_metadata(self, workbook):
        """Reapply comments/font to the latest workbook after 3-way value merge."""
        if not self._coa_commit_metadata:
            return
        worksheet = next(
            (sheet for sheet in workbook.Worksheets
             if str(sheet.Name).strip().lower() not in {"db", "guide", "가이드", "가이드(guide)", "index", "old chemical list"}),
            None,
        )
        if worksheet is None:
            return
        header_map, _ = build_header_map(worksheet)
        if not all(header in header_map for header in ("Order No.", "COA Link", "COA Local Path")):
            return
        order_rows = {}
        last_row = max(1, worksheet.Cells(worksheet.Rows.Count, header_map["Order No."]).End(-4162).Row)
        for row in range(2, last_row + 1):
            order = self.as_excel_text(worksheet.Cells(row, header_map["Order No."]).Value)
            if order and order not in order_rows:
                order_rows[order] = row
        for order, payload in self._coa_commit_metadata:
            row = order_rows.get(order)
            if row:
                COAManager.apply_metadata(worksheet, row, header_map, payload)

    def _commit_working_copy(self, target_path, changed, force_metadata=False):
        """Merge a completed private copy into the latest shared workbook."""
        if not changed:
            return 0

        is_auto_sync, max_retries = self._auto_sync_limits()
        if not check_and_wait_lock(
            target_path, log_fn=self.log, max_retries=max_retries,
            retry_delay=60, is_auto_sync=is_auto_sync, check_stop_fn=self.is_stopped,
        ):
            raise Exception("ChemicalList.xlsx 파일 사용 대기가 중단되었거나 다음 자동 동기화 주기까지 잠겨 있습니다.")
        if self.is_stopped():
            raise Exception("사용자에 의해 동기화가 중단되었습니다.")

        latest_fingerprint = file_fingerprint(target_path)
        latest_snapshot = snapshot_workbook(target_path)
        our_snapshot = snapshot_workbook(self._working_copy_path)
        patches, conflicts = build_three_way_plan(
            self._base_snapshot, our_snapshot, latest_snapshot
        )
        if conflicts:
            raise ConcurrentEditConflict(conflicts)
        if not patches and not force_metadata:
            self.log("공동편집 최신본과 비교한 결과 실제 반영할 셀이 없습니다.")
            return 0

        # Recheck immediately before opening the writable commit session.  A local
        # OneDrive update arriving during diff calculation invalidates the plan.
        if file_fingerprint(target_path) != latest_fingerprint:
            raise ConcurrentEditConflict(["저장 직전 OneDrive 최신본이 다시 변경됨"])

        commit_excel = None
        commit_wb = None
        try:
            commit_excel = win32com.client.DispatchEx("Excel.Application")
            commit_excel.Visible = False
            commit_excel.DisplayAlerts = False
            commit_excel.ScreenUpdating = False
            commit_excel.EnableEvents = False
            try:
                commit_excel.Calculation = -4135  # xlCalculationManual
            except Exception:
                pass
            commit_wb = commit_excel.Workbooks.Open(target_path, UpdateLinks=False, ReadOnly=False)
            if commit_wb.ReadOnly:
                raise Exception("ChemicalList.xlsx 최신본이 읽기 전용으로 열려 저장할 수 없습니다.")
            if file_fingerprint(target_path) != latest_fingerprint:
                raise ConcurrentEditConflict(["쓰기 세션을 여는 동안 OneDrive 최신본이 변경됨"])
            self._apply_cell_patches(commit_wb, patches)
            self._apply_coa_commit_metadata(commit_wb)

            # Validation and conditional-format coverage are program-managed
            # metadata.  Extend them without reinjecting broad cell formulas or
            # changing user values/alignment in the latest shared workbook.
            commit_main_sheet = None
            for sheet in commit_wb.Worksheets:
                lowered = str(sheet.Name).strip().lower()
                if lowered == "db":
                    db_last_row = max(2, sheet.Cells(sheet.Rows.Count, 1).End(-4162).Row)
                    self.apply_db_formatting(sheet, db_last_row)
                    continue
                if lowered in {"guide", "가이드", "가이드(guide)", "index", "old chemical list"}:
                    continue
                commit_main_sheet = sheet
                header_map = {}
                last_col = sheet.Cells(1, sheet.Columns.Count).End(-4159).Column
                for column in range(1, max(last_col + 1, 30)):
                    value = sheet.Cells(1, column).Value
                    if value:
                        header_map[str(value).strip()] = column
                # Formula patches from the private copy may contain stale
                # physical row references after a logical-key merge.
                self._apply_db_lookup_formulas(sheet)
                data_last_row = max(
                    2, meaningful_data_last_row(sheet, header_map, "order")
                )
                self.apply_formatting_and_validation(
                    sheet,
                    header_map,
                    format_validation_max_row(data_last_row),
                    write_formulas=False,
                    apply_layout=False,
                )
                break

            if commit_main_sheet is not None:
                commit_main_sheet.Activate()

            # COM edits are still only in memory, so a changed disk fingerprint
            # here means a coauthor/OneDrive update arrived during our commit.
            if file_fingerprint(target_path) != latest_fingerprint:
                raise ConcurrentEditConflict(["저장 직전 OneDrive 최신본이 변경됨"])
            if self.is_stopped():
                raise Exception("사용자에 의해 저장 작업이 중단되었습니다.")

            backup_path = create_backup(target_path)
            backup_name = os.path.basename(backup_path)
            self.log(f"변경사항 반영 전 최신 공유본을 백업했습니다: {backup_name}")

            commit_excel.ScreenUpdating = True
            commit_wb.Save()
            self.log(
                f"공동편집 최신본에 비충돌 변경 {len(patches)}개 셀과 "
                "프로그램 관리 서식을 병합했습니다."
            )
            return len(patches)
        finally:
            if commit_wb:
                try:
                    commit_wb.Close(SaveChanges=False)
                except Exception:
                    pass
            if commit_excel:
                try:
                    commit_excel.EnableEvents = True
                except Exception:
                    pass
                try:
                    commit_excel.Calculation = -4105  # xlCalculationAutomatic
                except Exception:
                    pass
                try:
                    commit_excel.Quit()
                except Exception:
                    pass

    def run_sync(self):
        self._coa_commit_metadata = []
        src_path = normalize_local_path(self.config.get("source_file", ""))
        
        src_sheet_name = self.config.get("source_sheet", "")
        header_row = int(self.config.get("header_row", 1))

        if not os.path.exists(src_path):
            raise Exception("원본 파일을 찾을 수 없습니다.")

        src_folder = os.path.dirname(src_path)
        
        target_path = resolve_target_file(self.config)
        self.target_file = target_path

        if os.path.normcase(src_path) == os.path.normcase(target_path):
            raise Exception("원본 오더북과 ChemicalList 대상 파일은 서로 달라야 합니다.")
        if self.config.get("target_file") and os.path.exists(target_path):
            valid_target, target_error = validate_chemical_list_file(target_path)
            if not valid_target:
                raise Exception(f"선택한 ChemicalList 파일이 올바르지 않습니다: {target_error}")

        deleted_backups, cleanup_failures = purge_expired_backups(months=3)
        if deleted_backups:
            self.log(f"3개월이 지난 백업 파일 {len(deleted_backups)}개를 삭제했습니다.")
        for failed_path, error in cleanup_failures:
            self.log(f"오래된 백업 파일 삭제 경고: {os.path.basename(failed_path)} ({error})")

        try:
            self.excel = win32com.client.DispatchEx("Excel.Application")
            self.excel.Visible = False
            self.excel.ScreenUpdating = False
            self.excel.DisplayAlerts = False
            
            self.log("원본 파일 여는 중 (읽기 전용 모드)...")
            self.source_wb = self.excel.Workbooks.Open(src_path, ReadOnly=True, UpdateLinks=False)
            
            src_ws = None
            for sheet in self.source_wb.Worksheets:
                if sheet.Name == src_sheet_name:
                    src_ws = sheet
                    break
            
            if not src_ws:
                raise Exception(f"'{src_sheet_name}' 시트를 찾을 수 없습니다.")

            # Target Setup
            is_new_target = not (target_path and os.path.exists(target_path))
            self.log("최신 ChemicalList 파일 확인 중...")
            if target_path and os.path.exists(target_path):
                is_auto_sync, max_retries = self._auto_sync_limits()
                 
                if not check_and_wait_lock(target_path, log_fn=self.log, max_retries=max_retries, retry_delay=60, is_auto_sync=is_auto_sync, check_stop_fn=self.is_stopped):
                    raise Exception("ChemicalList.xlsx 파일 사용 대기가 중단되었거나 파일이 열려있습니다.")

                work_path = self._make_working_copy(target_path)
                self.log("공동편집 기준본을 캡처하고 OneDrive 외부 작업 복사본을 열었습니다.")
                self.target_wb = self.excel.Workbooks.Open(work_path, UpdateLinks=False, ReadOnly=False)
                tgt_ws = self.target_wb.Sheets(1)
            else:
                self.log("ChemicalList 파일이 없어 새로 생성합니다.")
                fd, work_path = tempfile.mkstemp(prefix="ChemicalList_work_", suffix=".xlsx")
                os.close(fd)
                os.remove(work_path)
                self._working_copy_path = work_path
                self.target_wb = self.excel.Workbooks.Add()
                tgt_ws = self.target_wb.Sheets(1)
                tgt_ws.Name = "ChemicalList"

            # Build Target Indices dynamically from actual sheet row 1
            tgt_headers = self.config["target_headers"]
            tgt_col_map, tgt_last_col = build_header_map(tgt_ws)
            tgt_col_map = reserve_missing_headers(tgt_col_map, tgt_headers, tgt_last_col)
            
            # Ensure hidden Original Product Name column exists
            orig_pn_col = tgt_col_map.get("Original Product Name", 0)
            if orig_pn_col == 0:
                orig_pn_col = max(tgt_col_map.values()) + 1 if tgt_col_map else 16
                tgt_col_map["Original Product Name"] = orig_pn_col
            
            tgt_order_num_col = tgt_col_map.get("Order No.", 0)
            if tgt_order_num_col == 0:
                raise Exception("Target 헤더에 'Order No.' 가 없습니다.")

            # Find Source Columns dynamically based on row
            src_last_col = src_ws.Cells(header_row, src_ws.Columns.Count).End(-4159).Column # xlToLeft
            src_col_map = {}
            for c in range(1, src_last_col + 1):
                val = str(src_ws.Cells(header_row, c).Value).strip()
                if val:
                    src_col_map[val] = c
            
            src_order_num_col = src_col_map.get(self.config["mapping"].get("Order No.", ""), 0)
            if src_order_num_col == 0:
                raise Exception("원본 파일에서 주문 번호 매핑 열을 찾을 수 없습니다.")

            src_last_row = src_ws.Cells(src_ws.Rows.Count, src_order_num_col).End(-4162).Row # xlUp
            abs_src_last_row = src_ws.Cells.Find("*", SearchOrder=1, SearchDirection=2)
            if abs_src_last_row:
                src_last_row = max(src_last_row, abs_src_last_row.Row)
                
            if src_last_row <= header_row:
                raise Exception("원본 시트에 데이터가 없습니다.")

            # Read Source Data
            self.log("원본 데이터 읽기 중...")
            src_data = src_ws.Range(src_ws.Cells(header_row + 1, 1), src_ws.Cells(src_last_row, src_last_col)).Value
            
            if src_data is None:
                src_data = []
            elif not isinstance(src_data[0], (tuple, list)):
                src_data = [src_data]
            
            # Read Target Data
            tgt_prod_col = tgt_col_map.get("Product Name", 0)
            if tgt_prod_col > 0:
                tgt_last_row = tgt_ws.Cells(tgt_ws.Rows.Count, tgt_prod_col).End(-4162).Row
            else:
                tgt_last_row = tgt_ws.Cells(tgt_ws.Rows.Count, tgt_order_num_col).End(-4162).Row

            tgt_dict = {}
            if tgt_last_row >= 2:
                for r in range(2, tgt_last_row + 1):
                    val = str(tgt_ws.Cells(r, tgt_order_num_col).Value).strip()
                    if val and val != "None":
                        if val not in tgt_dict:
                            tgt_dict[val] = r

            mapping = self.config["mapping"]
            
            cnt_new = 0
            cnt_upd = 0
            cnt_dup = 0
            cnt_db_upd = 0

            from core.db_manager import DBManager
            total_updates = 0
            for row_data in src_data:
                if not row_data: continue
                order_num_str = str(row_data[src_order_num_col - 1]).strip() if src_order_num_col - 1 < len(row_data) else ""
                if not order_num_str or order_num_str == "None": continue
                if order_num_str not in tgt_dict:
                    total_updates += 1
            fast_mode = total_updates <= 5

            # db_manager는 target_wb(즉, ChemicalList.xlsx)에 쓰거나 읽을 때 열어둔 파일과 충돌날 수 있습니다.
            # 하지만 openpyxl은 별도로 열고 닫으므로, 엑셀 앱에서 열기 전에 미리 처리하거나 
            # 일단 여기서는 pandas 읽기(load_db)만 하고 저장은 다른 곳에 위임하는 식으로 해야합니다.
            # win32com에서 이미 열었기 때문에 openpyxl에서 쓰기 권한이 막힐 수 있습니다.
            # 이를 해결하기 위해 win32com 엑셀 파일은 임시로 생성하거나 다루고, 
            # DB 시트는 win32com 객체를 이용해 읽고 쓰는 방식으로 해야 충돌이 나지 않습니다.
            
            self.log("데이터 동기화 진행 중...")
            
            # Find specific cols for logic
            qty_col = tgt_col_map.get("Quantity", 0)
            used_col = tgt_col_map.get("Used", 0)
            stat_col = tgt_col_map.get("Status", 0)
            
            receipt_src_col = src_col_map.get("수령확인", 0)
            chem_src_col = src_col_map.get("품목명", src_col_map.get("시약", 0))

            # Network phase: derive the post-sync product set without changing
            # any Excel cell, then crawl each missing logical product once.
            pre_mfr_col = SyncEngine.get_col_idx(["Manufacturer", "제조사", "회사"], tgt_col_map)
            pre_cat_col = SyncEngine.get_col_idx(["Catalog No.", "제품번호", "품번", "카탈로그 번호"], tgt_col_map)
            candidate_products = {}
            for row_num in range(2, tgt_last_row + 1):
                pre_man = DBManager.normalize_manufacturer(self.as_excel_text(tgt_ws.Cells(row_num, pre_mfr_col).Value)) if pre_mfr_col else ""
                pre_cat = self.as_excel_text(tgt_ws.Cells(row_num, pre_cat_col).Value) if pre_cat_col else ""
                if pre_man and pre_cat:
                    candidate_products.setdefault(DBManager.crawl_key(pre_man, pre_cat), (pre_man, pre_cat))

            src_mfr_name = mapping.get("Manufacturer")
            src_cat_name = mapping.get("Catalog No.")
            src_mfr_col = src_col_map.get(src_mfr_name, 0) if src_mfr_name else 0
            src_cat_col = src_col_map.get(src_cat_name, 0) if src_cat_name else 0
            for row_data in src_data:
                if not row_data or not receipt_src_col or not self.is_received(row_data[receipt_src_col - 1]):
                    continue
                pre_man = DBManager.normalize_manufacturer(self.as_excel_text(row_data[src_mfr_col - 1])) if src_mfr_col else ""
                pre_cat = self.as_excel_text(row_data[src_cat_col - 1]) if src_cat_col else ""
                if pre_man and pre_cat:
                    candidate_products.setdefault(DBManager.crawl_key(pre_man, pre_cat), (pre_man, pre_cat))

            existing_crawl_keys = set()
            existing_db_sheet = next((sheet for sheet in self.target_wb.Worksheets if sheet.Name == "DB"), None)
            if existing_db_sheet:
                existing_headers, _ = build_header_map(existing_db_sheet)
                existing_db_last = existing_db_sheet.Cells(existing_db_sheet.Rows.Count, 1).End(-4162).Row
                for row_num in range(2, existing_db_last + 1):
                    existing_record = {
                        header: existing_db_sheet.Cells(row_num, column).Value
                        for header, column in existing_headers.items()
                    }
                    if not DBManager.needs_recrawl(existing_record):
                        existing_crawl_keys.add(DBManager.crawl_key(
                            existing_record.get("Manufacturer"),
                            existing_record.get("Catalog No."),
                        ))

            crawl_results = {}
            crawl_requests = {key: value for key, value in candidate_products.items() if key not in existing_crawl_keys}
            crawl_total = len(crawl_requests)
            for crawl_index, (key, (pre_man, pre_cat)) in enumerate(crawl_requests.items(), 1):
                if self.is_stopped():
                    raise Exception("사용자 요청으로 동기화가 중단되었습니다.")
                self.log(
                    f"DB 정보 보완 중 ({crawl_index}/{crawl_total}): "
                    f"{pre_man} {pre_cat}"
                )
                from scrapers.registry import create_scraper, scraper_class
                if scraper_class(pre_man) is None:
                    crawl_results[key] = {"error": "Manual Entry Required"}
                    continue
                if not hasattr(self, 'sb_context_manager') or not self.sb_context_manager:
                    from seleniumbase import SB
                    self.sb_context_manager = SB(uc=True, headless=self.config.get("headless", True))
                    self.sb = self.sb_context_manager.__enter__()
                try:
                    scraper = create_scraper(
                        pre_man, browser_context=self.sb, fast_mode=fast_mode,
                        base_dir=os.path.dirname(self.target_file), check_stop_fn=self.is_stopped,
                    )
                    crawl_results[key] = scraper.scrape(pre_cat) if scraper else {"error": "Manual Entry Required"}
                except Exception as error:
                    self.log(f"크롤링 경고 [{pre_man} {pre_cat}]: {error}")
                    crawl_results[key] = {"error": f"Scraping error: {error}"}
            if crawl_requests:
                target_labels = ", ".join(
                    f"{manufacturer} {catalog}"
                    for manufacturer, catalog in crawl_requests.values()
                )
                self.log(
                    f"DB 정보 보완 대상 {len(crawl_results)}건 조회 완료 "
                    f"[{target_labels}] (신규 주문 건수와는 별도). "
                    "Excel 반영을 시작합니다."
                )

            # COA phase is also completed before the first Excel write.  The
            # portable downloader remains untouched and is called once per
            # unique manufacturer/catalog/lot combination.
            coa_manager = COAManager(
                os.path.join(os.path.dirname(self.target_file), "coa"),
                log_fn=self.log,
                stop_fn=self.is_stopped,
            )
            coa_requests_by_order = {}
            coa_download_requests = []
            src_lot_name = mapping.get("Lot No.", "Lot No.")
            src_lot_col = src_col_map.get(src_lot_name, 0)
            coa_path_col = tgt_col_map.get("COA Local Path", 0)
            tgt_lot_col = tgt_col_map.get("Lot No.", 0)
            if src_lot_col and src_mfr_col and src_cat_col:
                for row_data in src_data:
                    if not row_data or not receipt_src_col or not self.is_received(row_data[receipt_src_col - 1]):
                        continue
                    order_number = self.as_excel_text(row_data[src_order_num_col - 1])
                    vendor = DBManager.normalize_manufacturer(self.as_excel_text(row_data[src_mfr_col - 1]))
                    catalog = self.as_excel_text(row_data[src_cat_col - 1])
                    lot = normalize_lot(row_data[src_lot_col - 1])
                    if not order_number or not vendor or not catalog or not lot or not is_supported_vendor(vendor):
                        continue
                    request = {"order": order_number, "vendor": vendor, "catalog": catalog, "lot": lot}
                    coa_requests_by_order[order_number] = request
                    existing_row = tgt_dict.get(order_number)
                    cached = False
                    if existing_row and coa_path_col and tgt_lot_col:
                        old_lot = normalize_lot(tgt_ws.Cells(existing_row, tgt_lot_col).Value)
                        old_path = tgt_ws.Cells(existing_row, coa_path_col).Value
                        cached = lots_equal(old_lot, lot) and valid_cached_document(old_path, catalog, lot)
                    if not cached:
                        coa_download_requests.append(request)

            coa_results = {}
            if coa_download_requests:
                if not hasattr(self, 'sb_context_manager') or not self.sb_context_manager:
                    from seleniumbase import SB
                    self.sb_context_manager = SB(uc=True, headless=self.config.get("headless", True))
                    self.sb = self.sb_context_manager.__enter__()
                self.log(f"COA/CoC 다운로드 대상 {len({coa_manager.key(r['vendor'], r['catalog'], r['lot']) for r in coa_download_requests})}건을 처리합니다.")
                coa_results = coa_manager.download_many(self.sb, coa_download_requests)

            # Excel write phase begins here.  Header repair, row updates and DB
            # writes happen only after every browser/network request has ended.
            write_reserved_headers(tgt_ws, tgt_col_map)
            c_dict = self.config.get("colors", {})
            hrange = tgt_ws.Range(tgt_ws.Cells(1, 1), tgt_ws.Cells(1, max(tgt_col_map.values())))
            hrange.Font.Bold = True
            hrange.Font.Color = hex_to_bgr(c_dict.get("header_font", "#ffffff"))
            hrange.Interior.Color = hex_to_bgr(c_dict.get("header_bg", "#464646"))
            hrange.HorizontalAlignment = -4108
            tgt_ws.Columns(orig_pn_col).Hidden = True

            for i, row_data in enumerate(src_data):
                if not row_data:
                    continue
                
                # src_data is 0-indexed tuple of tuples
                order_num_val = row_data[src_order_num_col - 1] if src_order_num_col - 1 < len(row_data) else None
                if not order_num_val:
                    continue
                order_num_str = str(order_num_val).strip()
                if not order_num_str or order_num_str == "None":
                    continue

                # Check if this row meets inclusion criteria
                if receipt_src_col == 0:
                    continue  # "수령확인" 열을 찾을 수 없으면 해당 행 스킵
                receipt_val = row_data[receipt_src_col - 1]
                if not self.is_received(receipt_val):
                    continue # Not received yet
                
                # Prepare identifier for logging
                chem_name = str(row_data[chem_src_col - 1]).strip() if chem_src_col > 0 and row_data[chem_src_col - 1] is not None else "Unknown"
                identifier = f"[{chem_name}] (Order: {order_num_str})"

                exists = order_num_str in tgt_dict
                
                if exists:
                    tgt_r = tgt_dict[order_num_str]
                    changed_cols = []
                    for t_header, s_header in mapping.items():
                        if not s_header: continue
                        if t_header == "Status": continue # Status is handled by formula
                        t_c = tgt_col_map.get(t_header, 0)
                        s_c = src_col_map.get(s_header, 0)
                        if t_c > 0 and s_c > 0 and s_c - 1 < len(row_data):
                            new_val = row_data[s_c - 1]
                            if t_header in self.TEXT_FIELDS:
                                new_val = self.as_excel_text(new_val)
                            if new_val is None: new_val = ""
                            old_val = tgt_ws.Cells(tgt_r, t_c).Value
                            if old_val is None: old_val = ""
                            
                            # 기존 값이 이미 적혀있다면 덮어쓰지 않음
                            if str(old_val).strip() != "" and str(old_val).strip() != "None":
                                continue
                                
                            if str(old_val).strip() != str(new_val).strip() and str(new_val).strip() != "":
                                changed_cols.append(f"{t_header}: 빈칸 -> '{str(new_val).strip()}'")
                                if t_header in self.TEXT_FIELDS:
                                    tgt_ws.Cells(tgt_r, t_c).NumberFormat = "@"
                                tgt_ws.Cells(tgt_r, t_c).Value = new_val
                                
                    if changed_cols:
                        self.log(f"  [업데이트] {identifier} - 변경항목: " + ", ".join(changed_cols))
                        cnt_upd += 1
                    else:
                        cnt_dup += 1
                else:
                    self.log(f"  [신규추가] {identifier}")
                    tgt_last_row += 1
                    tgt_r = tgt_last_row
                    tgt_dict[order_num_str] = tgt_r
                    
                    for t_header, s_header in mapping.items():
                        if not s_header: continue
                        if t_header == "Status": continue
                        t_c = tgt_col_map.get(t_header, 0)
                        s_c = src_col_map.get(s_header, 0)
                        if t_c > 0 and s_c > 0 and s_c - 1 < len(row_data):
                            new_val = row_data[s_c - 1]
                            if new_val is not None:
                                if t_header in self.TEXT_FIELDS:
                                    new_val = self.as_excel_text(new_val)
                                    tgt_ws.Cells(tgt_r, t_c).NumberFormat = "@"
                                tgt_ws.Cells(tgt_r, t_c).Value = new_val
                                
                tgt_ws.Range(tgt_ws.Cells(tgt_r, 1), tgt_ws.Cells(tgt_r, max(tgt_col_map.values()))).HorizontalAlignment = -4108
                coa_request = coa_requests_by_order.get(order_num_str)
                if coa_request:
                    coa_key = coa_manager.key(coa_request["vendor"], coa_request["catalog"], coa_request["lot"])
                    coa_payload = coa_manager.payload(coa_results.get(coa_key))
                    if coa_payload:
                        coa_values = {
                            "Expiration Date": coa_payload["expiration"],
                            "COA Link": coa_payload["link"],
                            "COA Local Path": coa_payload["path"],
                        }
                        coa_changed = False
                        for coa_header, coa_value in coa_values.items():
                            cell = tgt_ws.Cells(tgt_r, tgt_col_map[coa_header])
                            if str(cell.Value or "").strip() != str(coa_value or "").strip():
                                cell.Value = coa_value
                                coa_changed = True
                        coa_manager.apply_metadata(tgt_ws, tgt_r, tgt_col_map, coa_payload)
                        self._coa_commit_metadata.append((order_num_str, coa_payload))
                        if exists and coa_changed and not changed_cols:
                            cnt_upd += 1
                if not exists:
                    cnt_new += 1

            self.log("Chemical List 내 전체 시약 정보 크롤링 및 DB 동기화 검토 중...")
            
            db_sheet_name = "DB"
            db_ws = None
            for sheet in self.target_wb.Worksheets:
                if sheet.Name == db_sheet_name:
                    db_ws = sheet
                    break
            
            if not db_ws:
                db_ws = self.target_wb.Worksheets.Add(After=self.target_wb.Worksheets(self.target_wb.Worksheets.Count))
                db_ws.Name = db_sheet_name
                db_cols = ["Key", "Manufacturer", "Catalog No.", "Product Name", "CAS No.", "Storage Temp.", "Sensitivity", "Signal Word", "Key Hazards", "Detailed Hazard Classification", "Detail_Link", "SDS_Link", "SDS_Local_Path", "Revision Date"]
                for c_idx, c_name in enumerate(db_cols, 1):
                    db_ws.Cells(1, c_idx).Value = c_name
            
            db_last_row = db_ws.Cells(db_ws.Rows.Count, 1).End(-4162).Row
            
            db_cols_idx = {}
            for c in range(1, 20):
                val = db_ws.Cells(1, c).Value
                if val: db_cols_idx[str(val).strip()] = c
            db_last_col = max(db_cols_idx.values(), default=0)
            for header in DBManager.COLUMNS:
                if header not in db_cols_idx:
                    db_last_col += 1
                    db_ws.Cells(1, db_last_col).Value = header
                    db_cols_idx[header] = db_last_col
            db_last_row = meaningful_data_last_row(db_ws, db_cols_idx, "db")
                
            from core.db_manager import DBManager
            
            def get_tc(names):
                return SyncEngine.get_col_idx(names, tgt_col_map)

            def get_db_col(names):
                return SyncEngine.get_col_idx(names, db_cols_idx)

            if "Key" not in db_cols_idx:
                db_ws.Columns(1).Insert()
                db_ws.Cells(1, 1).Value = "Key"
                db_cols_idx = {}
                for c in range(1, 20):
                    val = db_ws.Cells(1, c).Value
                    if val: db_cols_idx[str(val).strip()] = c

            m_c = get_db_col(["Manufacturer", "제조사"])
            cat_c = get_db_col(["Catalog No.", "제품번호"])
            rev_c = get_db_col(["Revision Date", "갱신일"])
            k_db_col = get_db_col(["Key"])
            key_mfr_letter = get_col_letter(m_c)
            key_cat_letter = get_col_letter(cat_c)

            for r in range(2, db_last_row + 1):
                db_ws.Cells(r, k_db_col).Formula = (
                    f'={key_mfr_letter}{r}&"|"&{key_cat_letter}{r}'
                )
                if m_c > 0:
                    db_ws.Cells(r, m_c).NumberFormat = "@"
                    self.safe_set_val(db_ws.Cells(r, m_c), self.as_excel_text(db_ws.Cells(r, m_c).Value))
                if cat_c > 0:
                    db_ws.Cells(r, cat_c).NumberFormat = "@"
                    self.safe_set_val(db_ws.Cells(r, cat_c), self.as_excel_text(db_ws.Cells(r, cat_c).Value))
                if rev_c > 0:
                    revision_cell = db_ws.Cells(r, rev_c)
                    r_val = DBManager.normalize_revision_date(revision_cell.Value2)
                    revision_cell.NumberFormat = "@"
                    if r_val:
                        revision_cell.Value = r_val

            for tgt_r in range(2, tgt_last_row + 1):
                mfr_col = get_tc(["Manufacturer", "제조사", "회사"])
                cat_col = get_tc(["Catalog No.", "제품번호", "품번", "카탈로그 번호"])
                name_col = get_tc(["Product Name", "시약명", "품목명", "제품명"])
                
                raw_m = self.as_excel_text(tgt_ws.Cells(tgt_r, mfr_col).Value) if mfr_col > 0 else ""
                product_num = self.as_excel_text(tgt_ws.Cells(tgt_r, cat_col).Value) if cat_col > 0 else ""
                
                manufacturer = DBManager.normalize_manufacturer(raw_m)
                if manufacturer and mfr_col > 0:
                    tgt_ws.Cells(tgt_r, mfr_col).NumberFormat = "@"
                    self.safe_set_val(tgt_ws.Cells(tgt_r, mfr_col), manufacturer)
                    
                product_num = product_num if product_num and product_num != "None" else ""
                    
                if cat_col > 0 and product_num:
                    tgt_ws.Cells(tgt_r, cat_col).NumberFormat = "@"
                    self.safe_set_val(tgt_ws.Cells(tgt_r, cat_col), product_num)
                    
                chem_name_fallback = tgt_ws.Cells(tgt_r, name_col).Value if name_col > 0 else "Unknown"
                chem_name_fallback = str(chem_name_fallback).strip() if chem_name_fallback and chem_name_fallback != "None" else "Unknown"
                
                if manufacturer and product_num:
                    db_result = None
                    existing_db_row = 0
                    k_db_col = get_db_col(["Key"])
                    m_db_col = get_db_col(["Manufacturer", "제조사"])
                    c_db_col = get_db_col(["Catalog No.", "제품번호"])
                    
                    target_key = f"{manufacturer}|{product_num}"
                    
                    for r in range(2, db_last_row + 1):
                        r_key = str(db_ws.Cells(r, k_db_col if k_db_col > 0 else 1).Value).strip()
                        r_man = DBManager.normalize_manufacturer(db_ws.Cells(r, m_db_col if m_db_col > 0 else 2).Value)
                        r_num = str(db_ws.Cells(r, c_db_col if c_db_col > 0 else 3).Value).strip()
                        if r_num.endswith(".0"): r_num = r_num[:-2]
                        
                        if (r_key and r_key.lower() == target_key.lower()) or (r_man.lower() == manufacturer.lower() and r_num == product_num):
                            existing_db_row = r
                            db_result = {}
                            for k, idx in db_cols_idx.items():
                                db_result[k] = db_ws.Cells(r, idx).Value
                            if DBManager.needs_recrawl(db_result):
                                db_result = None
                            break
                            
                    if not db_result:
                        crawled_data = crawl_results.get(
                            DBManager.crawl_key(manufacturer, product_num),
                            {"error": "Missing prefetched scraping result"},
                        )
                        
                        if crawled_data:
                            if "error" in crawled_data:
                                is_man_req = crawled_data["error"] in ["Manual Entry Required", "Manual Input Required"]
                                db_result = {
                                    "Manufacturer": manufacturer,
                                    "Catalog No.": product_num,
                                    "Product Name": chem_name_fallback,
                                    "CAS No.": "Manual Input Required" if is_man_req else "Search Failed",
                                    "Storage Temp.": "-",
                                    "Signal Word": "-",
                                    "Key Hazards": "-",
                                    "Detailed Hazard Classification": "-",
                                    "Sensitivity": "-",
                                    "Detail_Link": "-" if is_man_req else "Product Not Found",
                                    "SDS_Link": "-",
                                    "SDS_Local_Path": "-",
                                    "Revision Date": "-"
                                }
                            else:
                                db_result = crawled_data
                                db_result["Manufacturer"] = manufacturer
                                db_result["Catalog No."] = product_num
                                
                                p_name = db_result.get("Product Name", db_result.get("시약명", ""))
                                if p_name in ["제조사 홈페이지에서 검색 실패", "검색 실패", "Search Failed", "Product Not Found", "", None] or str(p_name).strip() == "":
                                    db_result["Product Name"] = chem_name_fallback
                                    db_result["CAS No."] = "Search Failed"
                                    db_result["Detail_Link"] = "Product Not Found"
                                    db_result["Storage Temp."] = "-"
                                    db_result["Sensitivity"] = "-"
                                    db_result["Signal Word"] = "-"
                                    db_result["Key Hazards"] = "-"
                                    db_result["Detailed Hazard Classification"] = "-"
                                    db_result["SDS_Link"] = "-"
                                    db_result["SDS_Local_Path"] = "-"
                                    db_result["Revision Date"] = "-"
                                else:
                                    if db_result.get("CAS No.", db_result.get("CAS Number", "")) in ["정보 없음", "", "nan", "None", "N/A", "-"]:
                                        db_result["CAS No."] = "N/A"
                                        
                                    norm_temp = DBManager.normalize_temperature(db_result.get("Storage Temp.", db_result.get("보관온도", "")))
                                    db_result["Storage Temp."] = norm_temp
                                    
                                    det_haz = db_result.get("Detailed Hazard Classification", db_result.get("상세 위험분류", ""))
                                    sens = DBManager.extract_sensitivity(det_haz)
                                    db_result["Sensitivity"] = sens if sens else db_result.get("Sensitivity", "-")
                                    if not db_result.get("Signal Word"): db_result["Signal Word"] = "-"
                                    if not db_result.get("Key Hazards"): db_result["Key Hazards"] = "-"
                                    if not db_result.get("Detailed Hazard Classification"): db_result["Detailed Hazard Classification"] = "-"
                                    if not db_result.get("Detail_Link"): db_result["Detail_Link"] = "-"
                                    if not db_result.get("SDS_Link"): db_result["SDS_Link"] = "-"
                                    if not db_result.get("SDS_Local_Path"): db_result["SDS_Local_Path"] = "-"
                                    db_result["Revision Date"] = datetime.datetime.now().strftime("%Y-%m-%d")
                             
                            # Deduplicated DB writing
                            if existing_db_row > 0:
                                target_db_row = existing_db_row
                            else:
                                db_last_row += 1
                                target_db_row = db_last_row
                                
                            max_c = max(db_cols_idx.values()) if db_cols_idx else 14
                            
                            # Identifier columns must remain text even when their value is numeric.
                            c_mfr_idx = db_cols_idx.get("Manufacturer", 2)
                            c_cat_idx = db_cols_idx.get("Catalog No.", 3)
                            db_ws.Cells(target_db_row, c_mfr_idx).NumberFormat = "@"
                            db_ws.Cells(target_db_row, c_cat_idx).NumberFormat = "@"
                            db_ws.Cells(target_db_row, db_cols_idx.get("Revision Date", 14)).NumberFormat = "@"
                            
                            for k, v in db_result.items():
                                if k == "Key": continue
                                c_idx = db_cols_idx.get(k)
                                if c_idx and c_idx <= max_c:
                                    if k in self.TEXT_FIELDS:
                                        v = self.as_excel_text(v)
                                    self.safe_set_val(db_ws.Cells(target_db_row, c_idx), v)
                                    
                            # Inject Key Excel Formula in Col A
                            db_ws.Cells(target_db_row, k_db_col).Formula = (
                                f'={key_mfr_letter}{target_db_row}&"|"&'
                                f'{key_cat_letter}{target_db_row}'
                            )
                            cnt_db_upd += 1
                        else:
                            db_result = {
                                "Manufacturer": manufacturer, "Catalog No.": product_num, "Product Name": chem_name_fallback,
                                "CAS No.": "Search Failed", "Detail_Link": "Product Not Found",
                                "Storage Temp.": "-", "Signal Word": "-", "Key Hazards": "-", "Detailed Hazard Classification": "-",
                                "Sensitivity": "-", "SDS_Link": "-", "SDS_Local_Path": "-", "Revision Date": "-"
                            }
                            if existing_db_row > 0:
                                target_db_row = existing_db_row
                            else:
                                db_last_row += 1
                                target_db_row = db_last_row
                                
                            max_c = max(db_cols_idx.values()) if db_cols_idx else 14
                            c_mfr_idx = db_cols_idx.get("Manufacturer", 2)
                            c_cat_idx = db_cols_idx.get("Catalog No.", 3)
                            db_ws.Cells(target_db_row, c_mfr_idx).NumberFormat = "@"
                            db_ws.Cells(target_db_row, c_cat_idx).NumberFormat = "@"
                            db_ws.Cells(target_db_row, db_cols_idx.get("Revision Date", 14)).NumberFormat = "@"
                            
                            for k, v in db_result.items():
                                if k == "Key": continue
                                c_idx = db_cols_idx.get(k)
                                if c_idx and c_idx <= max_c:
                                    if k in self.TEXT_FIELDS:
                                        v = self.as_excel_text(v)
                                    self.safe_set_val(db_ws.Cells(target_db_row, c_idx), v)
                                    
                            db_ws.Cells(target_db_row, k_db_col).Formula = (
                                f'={key_mfr_letter}{target_db_row}&"|"&'
                                f'{key_cat_letter}{target_db_row}'
                            )
                            cnt_db_upd += 1
                            
                    if db_result:
                        from openpyxl.utils import get_column_letter
                        cat_col_let = get_column_letter(cat_col) if cat_col > 0 else "H"
                        mfr_col_let = get_column_letter(mfr_col) if mfr_col > 0 else "E"

                        lookup_map = [
                            # Storage Temp. is handled separately by the header-based formula pass.
                            (["Sensitivity", "민감성"], ["Sensitivity", "민감성"]),
                            (["Signal Word", "신호어"], ["Signal Word", "신호어"]),
                            (["Key Hazards", "주요위험"], ["Key Hazards", "Key Hazards", "주요위험", "Major Hazard"]),
                            (["Detailed Hazard Classification", "상세 위험분류"], ["Detailed Hazard Classification", "Detailed Hazard Classification", "상세 위험분류", "Hazard Statement"]),
                            (["Detail_Link", "상세정보_링크"], ["Detail_Link", "Detail_Link", "상세정보_링크", "Product Link"]),
                            (["SDS_Link"], ["SDS_Link", "SDS Link"]),
                            (["SDS_Local_Path"], ["SDS_Local_Path", "SDS Path"])
                        ]
                        
                        for db_keys, header_names in lookup_map:
                            t_c = get_tc(header_names)
                            if t_c > 0:
                                db_c = 0
                                for k in db_keys:
                                    if k in db_cols_idx:
                                        db_c = db_cols_idx[k]
                                        break
                                if db_c >= 1:
                                    db_value_let = get_column_letter(db_c)
                                    db_key_let = get_column_letter(db_cols_idx["Key"])
                                    lookup = f'INDEX(DB!${db_value_let}:${db_value_let},MATCH({mfr_col_let}{tgt_r}&"|"&{cat_col_let}{tgt_r},DB!${db_key_let}:${db_key_let},0))'
                                    formula = f'=IF(OR({mfr_col_let}{tgt_r}="", {cat_col_let}{tgt_r}=""), "", IFERROR(IF({lookup}="","-",{lookup}), "-"))'
                                    self.safe_set_val(tgt_ws.Cells(tgt_r, t_c), formula)

                        pname_tc = get_tc(["Product Name", "시약명", "품명", "제품명"])
                        orig_pn_tc = get_tc(["Original Product Name"])
                        row_order = tgt_ws.Cells(tgt_r, tgt_order_num_col).Value
                        current_product_formula = (
                            str(getattr(tgt_ws.Cells(tgt_r, pname_tc), "Formula", "") or "")
                            if pname_tc > 0 else ""
                        )
                        if (
                            pname_tc > 0 and orig_pn_tc > 0
                            and should_manage_product_name(row_order, current_product_formula)
                        ):
                            orig_val = tgt_ws.Cells(tgt_r, orig_pn_tc).Value
                            if not orig_val or str(orig_val).strip() in ["", "None"]:
                                curr_p = tgt_ws.Cells(tgt_r, pname_tc).Value
                                fallback_str = str(curr_p).strip() if curr_p and not str(curr_p).startswith("=") else chem_name_fallback
                                tgt_ws.Cells(tgt_r, orig_pn_tc).Value = fallback_str
                            orig_let = get_column_letter(orig_pn_tc)
                            db_pn_c = db_cols_idx.get("Product Name", 4)
                            db_pn_let = get_column_letter(db_pn_c)
                            db_key_let = get_column_letter(db_cols_idx["Key"])
                            lookup = f'INDEX(DB!${db_pn_let}:${db_pn_let},MATCH({mfr_col_let}{tgt_r}&"|"&{cat_col_let}{tgt_r},DB!${db_key_let}:${db_key_let},0))'
                            pn_formula = f'=IF(OR({mfr_col_let}{tgt_r}="", {cat_col_let}{tgt_r}=""), {orig_let}{tgt_r}, IFERROR(IF({lookup}="", {orig_let}{tgt_r}, {lookup}), {orig_let}{tgt_r}))'
                            self.safe_set_val(tgt_ws.Cells(tgt_r, pname_tc), pn_formula)

                        cas_tc = get_tc(["CAS No.", "CAS Number", "CAS 번호"])
                        if cas_tc > 0:
                            curr = tgt_ws.Cells(tgt_r, cas_tc).Value
                            curr_str = str(curr).strip() if curr is not None else ""
                            if not curr_str or curr_str in ["None", "nan", "N/A", "Search Failed", "Manual Input Required", "Manual Entry Required"]:
                                db_cas_c = db_cols_idx.get("CAS No.", 5)
                                db_cas_let = get_column_letter(db_cas_c)
                                db_key_let = get_column_letter(db_cols_idx["Key"])
                                lookup = f'INDEX(DB!${db_cas_let}:${db_cas_let},MATCH({mfr_col_let}{tgt_r}&"|"&{cat_col_let}{tgt_r},DB!${db_key_let}:${db_key_let},0))'
                                formula = f'=IF(OR({mfr_col_let}{tgt_r}="", {cat_col_let}{tgt_r}=""), "", IFERROR({lookup}, ""))'
                                self.safe_set_val(tgt_ws.Cells(tgt_r, cas_tc), formula)

            # If no new items were added and no items were updated, do NOT perform formatting, completion coloring, or saving
            if cnt_new == 0 and cnt_upd == 0 and cnt_db_upd == 0:
                required_row = format_validation_max_row(tgt_last_row)
                needs_extension = (
                    not is_new_target
                    and workbook_range_needs_extension(target_path, required_row)
                )
                needs_activation = (
                    not is_new_target
                    and workbook_main_sheet_needs_activation(target_path)
                )
                self.cleanup()
                if needs_extension or needs_activation:
                    metadata_changes = []
                    if needs_extension:
                        metadata_changes.append(f"관리 범위를 {required_row:,}행까지 확장")
                    if needs_activation:
                        metadata_changes.append("기본 열기 시트를 ChemicalList로 지정")
                    self.log("데이터 변경은 없지만 " + ", ".join(metadata_changes) + "합니다.")
                    self._commit_working_copy(
                        target_path, changed=True, force_metadata=True
                    )
                else:
                    self.log("동기화할 새로운 내용이 없습니다 (기존 파일 유지).")
                return {
                    "success": True,
                    "new": 0,
                    "updated": 0,
                    "duplicate": cnt_dup,
                    "no_changes": True,
                    "formatting_updated": needs_extension,
                    "active_sheet_updated": needs_activation,
                }

            # Apply Conditional Formatting & Validation on Target.  The range
            # intentionally leaves room for well over 1,000 future rows.
            try:
                working_format_row = format_validation_max_row(tgt_last_row)
                self.apply_formatting_and_validation(
                    tgt_ws, tgt_col_map, working_format_row
                )
                sheet_names = [s.Name for s in self.target_wb.Worksheets]
                if "DB" in sheet_names:
                    db_ws = self.target_wb.Worksheets("DB")
                    db_last_r = db_ws.Cells(db_ws.Rows.Count, 1).End(-4162).Row
                    if db_last_r >= 2:
                        self.apply_db_formatting(db_ws, db_last_r)
            except Exception as e:
                self.log(f"서식 주입 중 오류 (무시됨): {e}")
            
            # --- Insert Guide Sheet ---
            try:
                guide_sheet_name = "Guide"
                sheet_names = [s.Name for s in self.target_wb.Worksheets]
                if "가이드(Guide)" in sheet_names and "Guide" not in sheet_names:
                    try:
                        self.target_wb.Worksheets("가이드(Guide)").Name = "Guide"
                    except:
                        pass
                    sheet_names = [s.Name for s in self.target_wb.Worksheets]

                if guide_sheet_name not in sheet_names:
                    guide_ws = self.target_wb.Worksheets.Add(After=self.target_wb.Worksheets(self.target_wb.Worksheets.Count))
                    guide_ws.Name = guide_sheet_name
                    
                    guide_text_rows = [
                        ["[ Chemical Manager 프로그램 사용 방법 및 주의사항 ]"],
                        [""],
                        ["1. 기본 동작 원리 및 저장 기준"],
                        ["- 프로그램은 원본 Orderbook.xlsx (주문대장) 파일의 변경을 감지하여 데이터를 읽어옵니다."],
                        ["- 폴더 내에서 이름에 적힌 날짜/시간이 가장 최신인 ChemicalList 파일을 기준으로 삼습니다."],
                        ["- 업데이트 완료 시 덮어쓰기 대신 새로운 날짜/시간이 적힌 새 파일(ChemicalList_YYYYMMDD_HHMMSS)을 생성합니다."],
                        ["- 기존의 낡은 파일들은 안전하게 'Old Chemical List' 폴더로 자동 이동됩니다."],
                        [""],
                        ["2. 사용자 작성 시 유의사항 (Orderbook 및 Chemical List)"],
                        ["- 빈 줄 금지: 데이터 중간에 완전히 비어있는 줄이 있으면 데이터를 읽다가 중단될 수 있으므로 차례대로 기입하세요."],
                        ["- 동기화 기준: 시약명과 카탈로그 번호 등을 기준으로 기존에 등록된 시약인지 새 시약인지 판단합니다."],
                        ["- 수령 확인 (Status): 수령 후 Status 열에 'O' 또는 'ㅇ'을 입력하면 수령으로 간주되며, PDF 출력 필터링에 사용됩니다."],
                        ["- 드롭다운 선택: 캐비넷(Cabinet)은 Room과 Storage Temp.에 따라 동적으로 변하므로 잘못된 값을 억지로 쓰지 마세요."],
                        [""],
                        ["3. 프로그램 사용 주의사항"],
                        ["- 열린 파일 처리: Chemical List 엑셀 파일이 켜져 있어도 프로그램이 알아서 우회하여 새 파일을 생성합니다."],
                        ["- 단, 원본인 Orderbook 파일은 엑셀에서 저장을 완료해야만 프로그램이 변경 사항을 정확히 감지할 수 있습니다."],
                        ["- 자동 동기화 켜짐 상태에서는 Orderbook을 저장할 때마다 병합이 진행되므로, 수동 제어를 원하시면 정지 버튼을 누르세요."]
                    ]
                    
                    for r, row_data in enumerate(guide_text_rows, 1):
                        guide_ws.Cells(r, 1).Value = row_data[0]
                        if r in [3, 9, 15]: # Subtitles
                            guide_ws.Cells(r, 1).Font.Bold = True
                            
                    title_rng = guide_ws.Range("A1:G1")
                    title_rng.Merge()
                    title_rng.Font.Size = 14
                    title_rng.Font.Bold = True
                    title_rng.Interior.Color = 14211288
                    title_rng.HorizontalAlignment = -4108
                    
                    guide_ws.Columns(1).ColumnWidth = 100
            except Exception as e:
                self.log(f"가이드 시트 생성 중 오류 (무시됨): {e}")

            # 엑셀 파일 열 때 항상 ChemicalList 시트가 보이도록 활성화
            # Rebuild every program-owned DB formula from actual headers after
            # formatting so no temporary fixed-range formula reaches disk.
            self._apply_db_lookup_formulas(tgt_ws)
            tgt_ws.Activate()

            changed = cnt_new > 0 or cnt_upd > 0 or cnt_db_upd > 0
            self.log("OneDrive 공유본과 병합하기 위한 작업 복사본 저장 중...")
            if is_new_target:
                self.target_wb.SaveAs(self._working_copy_path)
            else:
                self.target_wb.Save()

            self.cleanup()

            if is_new_target:
                if os.path.exists(target_path):
                    raise ConcurrentEditConflict(["새 파일 생성 중 다른 사용자가 ChemicalList.xlsx를 생성함"])
                shutil.copy2(self._working_copy_path, target_path)
                self.log("새 ChemicalList.xlsx 파일을 생성했습니다.")
            else:
                self._commit_working_copy(target_path, changed)

            return {
                "success": True,
                "new": cnt_new,
                "updated": cnt_upd,
                "duplicate": cnt_dup
            }

        except ConcurrentEditConflict as e:
            self.cleanup()
            self._discard_working_copy()
            if not self.is_stopped() and self._restart_count < 2:
                self._restart_count += 1
                self.log(f"{e} 최신 공유본으로 처음부터 다시 동기화합니다. ({self._restart_count}/2)")
                return self.run_sync()
            return {"success": False, "error": str(e), "conflict": True}
        except Exception as e:
            import traceback
            traceback.print_exc()
            return {"success": False, "error": str(e)}
        finally:
            self.cleanup()
            self._discard_working_copy()

    def apply_formatting_and_validation(
        self, ws, tgt_col_map, max_row, write_formulas=True, apply_layout=True
    ):
        self.log("조건부 서식 및 데이터 유효성 검사 주입 중...")
        c_dict = self.config.get("colors", {})
        data_key_col = (
            tgt_col_map.get("Order No.", 0)
            or tgt_col_map.get("Product Name", 0)
            or 1
        )
        data_last_row = max(2, ws.Cells(ws.Rows.Count, data_key_col).End(-4162).Row)
        
        # 1. Conditional Formatting for missing or '-' values
        try:
            prod_col = tgt_col_map.get("Product Name", 0)
            prod_letter = get_col_letter(prod_col) if prod_col > 0 else "D"
            max_c = max(tgt_col_map.values()) if tgt_col_map else 20
            max_c_let = get_col_letter(max_c)
            rng_all = ws.Range(f"A2:{max_c_let}{max_row}")
            try:
                # Rebuild the program-managed rules once to avoid accumulating
                # duplicate rules every sync.
                rng_all.FormatConditions.Delete()
            except Exception:
                pass
             
            def apply_cf(col_name, bg_key, fg_key):
                col_idx = tgt_col_map.get(col_name, 0)
                if col_idx > 0:
                    try:
                        letter = get_col_letter(col_idx)
                        rng = ws.Range(f"{letter}2:{letter}{max_row}")
                        fc = rng.FormatConditions.Add(2, None, f'=AND(OR(TRIM({letter}2)="", TRIM({letter}2)="-"), TRIM({prod_letter}2)<>"")')
                        fc.Interior.Color = hex_to_bgr(c_dict.get(bg_key, "#ffffff"))
                        fc.Font.Color = hex_to_bgr(c_dict.get(fg_key, "#000000"))
                    except Exception as e:
                        pass

            apply_cf("CAS No.", "warn_cas_bg", "warn_cas_font")
            apply_cf("Catalog No.", "warn_num_bg", "warn_num_font")
            apply_cf("Room", "warn_loc_bg", "warn_loc_font")
            apply_cf("Storage Temp.", "warn_loc_bg", "warn_loc_font")
            apply_cf("Cabinet", "warn_loc_bg", "warn_loc_font")

            # Conditional Formatting for Search Failed & Manual Input Required
            try:
                sf_bg = hex_to_bgr(c_dict.get("search_failed_bg", "#ffff00"))
                sf_fg = hex_to_bgr(c_dict.get("search_failed_font", "#ff0000"))
                mi_bg = hex_to_bgr(c_dict.get("manual_input_bg", "#ffff00"))
                mi_fg = hex_to_bgr(c_dict.get("manual_input_font", "#ff0000"))

                fc_sf = rng_all.FormatConditions.Add(1, 3, "=\"Search Failed\"")
                fc_sf.Interior.Color = sf_bg
                fc_sf.Font.Color = sf_fg
                fc_sf.Font.Bold = True

                fc_pnf = rng_all.FormatConditions.Add(1, 3, "=\"Product Not Found\"")
                fc_pnf.Interior.Color = sf_bg
                fc_pnf.Font.Color = sf_fg
                fc_pnf.Font.Bold = True

                fc_mi = rng_all.FormatConditions.Add(1, 3, "=\"Manual Input Required\"")
                fc_mi.Interior.Color = mi_bg
                fc_mi.Font.Color = mi_fg
                fc_mi.Font.Bold = True

                fc_me = rng_all.FormatConditions.Add(1, 3, "=\"Manual Entry Required\"")
                fc_me.Interior.Color = mi_bg
                fc_me.Font.Color = mi_fg
                fc_me.Font.Bold = True
            except Exception as e:
                pass
        except Exception as e:
            self.log(f"조건부 서식 주입 중 경고 (무시됨): {e}")

        # 2. Data Validation for Room, Storage Temp., and Cabinet
        try:
            room_col = tgt_col_map.get("Room", 0)
            stemp_col = tgt_col_map.get("Storage Temp.", 0)
            cab_col = tgt_col_map.get("Cabinet", 0)
            mfr_col = tgt_col_map.get("Manufacturer", 0)
            cat_col = tgt_col_map.get("Catalog No.", 0)
            
            if room_col > 0:
                room_l = get_col_letter(room_col)
                rng_r = ws.Range(f"{room_l}2:{room_l}{max_row}")
                try:
                    rng_r.Validation.Delete()
                    rng_r.Validation.Add(3, 1, 1, "='index'!$A$2:$A$4")
                except: pass
                
            if stemp_col > 0:
                stemp_l = get_col_letter(stemp_col)
                rng_s = ws.Range(f"{stemp_l}2:{stemp_l}{max_row}")
                try:
                    rng_s.Validation.Delete()
                    rng_s.Validation.Add(3, 1, 1, "='index'!$C$2:$C$5")
                except: pass

                if write_formulas and mfr_col > 0 and cat_col > 0:
                    mfr_l = get_col_letter(mfr_col)
                    cat_l = get_col_letter(cat_col)
                    # Dynamically determine Storage Temp column in DB sheet from headers
                    try:
                        from utils.excel_utils import get_col_letter as gcl
                        db_ws_tmp = None
                        for _s in ws.Parent.Worksheets:
                            if _s.Name == "DB":
                                db_ws_tmp = _s
                                break
                        if db_ws_tmp:
                            db_header_map, _ = build_header_map(db_ws_tmp)
                            db_stemp_c = db_header_map.get("Storage Temp.", 0)
                            db_key_c = db_header_map.get("Key", 0)
                        else:
                            db_stemp_c = 0
                            db_key_c = 0
                    except:
                        db_stemp_c = 0
                        db_key_c = 0
                    try:
                        if not db_stemp_c or not db_key_c:
                            raise ValueError("DB Key/Storage Temp. 헤더가 없습니다.")
                        db_stemp_l = get_col_letter(db_stemp_c)
                        db_key_l = get_col_letter(db_key_c)
                        formula_rng = ws.Range(f"{stemp_l}2:{stemp_l}{data_last_row}")
                        vals_mfr = ws.Range(f"{mfr_l}2:{mfr_l}{data_last_row}").Value
                        vals_cat = ws.Range(f"{cat_l}2:{cat_l}{data_last_row}").Value
                        existing = formula_rng.Formula
                        
                        if not isinstance(vals_mfr, tuple):
                            vals_mfr = ((vals_mfr,),)
                            vals_cat = ((vals_cat,),)
                            existing = ((existing,),)
                            
                        new_formulas = []
                        for i in range(len(vals_mfr)):
                            m = vals_mfr[i][0]
                            c = vals_cat[i][0]
                            e = existing[i][0]
                            if m and str(m).strip() and c and str(c).strip():
                                r = i + 2
                                lookup = f'INDEX(DB!${db_stemp_l}:${db_stemp_l},MATCH({mfr_l}{r}&"|"&{cat_l}{r},DB!${db_key_l}:${db_key_l},0))'
                                f = f'=IFERROR(IF(OR({lookup}="", {lookup}="-"), "", {lookup}), "")'
                                new_formulas.append((f,))
                            else:
                                new_formulas.append((e,))
                                
                        if len(new_formulas) == 1:
                            formula_rng.Formula = new_formulas[0][0]
                        elif len(new_formulas) > 1:
                            formula_rng.Formula = new_formulas
                    except: pass
            if cab_col > 0 and room_col > 0 and stemp_col > 0:
                cab_l = get_col_letter(cab_col)
                room_l = get_col_letter(room_col)
                stemp_l = get_col_letter(stemp_col)
                rng_c = ws.Range(f"{cab_l}2:{cab_l}{max_row}")
                try:
                    rng_c.Validation.Delete()
                    formula = f'=IF(COUNTIF(index!$H:$H, {room_l}2&"_"&{stemp_l}2)=0, index!$G$100, OFFSET(index!$G$1, MATCH({room_l}2&"_"&{stemp_l}2, index!$H:$H, 0)-1, 0, COUNTIF(index!$H:$H, {room_l}2&"_"&{stemp_l}2), 1))'
                    rng_c.Validation.Add(3, 1, 1, formula)
                except: pass
        except Exception as e:
            self.log(f"유효성 검사 주입 중 경고 (무시됨): {e}")

        # 2. Apply Status Formula & Coloring
        try:
            stat_col = tgt_col_map.get("Status", 0)
            qty_col = tgt_col_map.get("Quantity", 0)
            used_col = tgt_col_map.get("Used", 0)
            
            if stat_col > 0 and qty_col > 0 and used_col > 0:
                stat_l = get_col_letter(stat_col)
                qty_l = get_col_letter(qty_col)
                used_l = get_col_letter(used_col)
                if write_formulas:
                    prod_col = tgt_col_map.get("Product Name", 0)
                    if prod_col > 0:
                        prod_l = get_col_letter(prod_col)
                        formula = f'=IF(TRIM({prod_l}2)="", "", IF(N({qty_l}2)-N({used_l}2)<=0, "X", N({qty_l}2)-N({used_l}2)))'
                    else:
                        formula = f'=IF(N({qty_l}2)-N({used_l}2)<=0, "X", N({qty_l}2)-N({used_l}2))'
                    formula_rng = ws.Range(f"{stat_l}2:{stat_l}{data_last_row}")
                    formula_rng.Formula = formula
                    formula_rng.HorizontalAlignment = -4108 # xlCenter
                 
                # Status "X" coloring
                try:
                    rng = ws.Range(f"{stat_l}2:{stat_l}{max_row}")
                    rng.FormatConditions.Delete()
                    fc_x = rng.FormatConditions.Add(1, 3, "=\"X\"") # xlCellValue, xlEqual
                    fc_x.Interior.Color = hex_to_bgr(c_dict.get("status_x_bg", "#ffe6e6"))
                    fc_x.Font.Color = hex_to_bgr(c_dict.get("status_x_font", "#ff0000"))
                    fc_x.Font.Bold = True
                except Exception as e:
                    pass
        except Exception as e:
            self.log(f"Status 수식 적용 중 경고 (무시됨): {e}")
            
        # 3. Symbol Colors for Signal Word and Key Hazards
        try:
            sig_col = SyncEngine.get_col_idx(["Signal Word", "신호어"], tgt_col_map)
            haz_col = SyncEngine.get_col_idx(["Key Hazards", "주요위험", "Major Hazard"], tgt_col_map)
            if apply_layout:
                self.apply_symbol_colors(ws, max_row, sig_col)
                self.apply_symbol_colors(ws, max_row, haz_col)
        except Exception as e:
            self.log(f"기호 색상 적용 중 경고 (무시됨): {e}")

        # 4. Vertical & Horizontal Alignments
        try:
            if not apply_layout:
                return
            center_cols = {
                "Order No.", "Order Date", "Ordered By", "Package Size", 
                "CAS No.", "Catalog No.", "Room", "Storage Temp.", "Cabinet", 
                "Quantity", "Used", "Status"
            }
            for col_name, c_idx in tgt_col_map.items():
                if c_idx > 0:
                    letter = get_col_letter(c_idx)
                    col_rng = ws.Range(f"{letter}2:{letter}{max_row}")
                    col_rng.VerticalAlignment = -4108 # xlVCenter (상하 가운데)
                    if col_name in center_cols:
                        col_rng.HorizontalAlignment = -4108 # xlHAlignCenter (좌우 가운데)
                    else:
                        col_rng.HorizontalAlignment = -4131 # xlHAlignLeft (좌우 왼쪽)

            max_c_idx = max(tgt_col_map.values()) if tgt_col_map else 20
            hdr_let = get_col_letter(max_c_idx)
            header_rng = ws.Range(f"A1:{hdr_let}1")
            header_rng.VerticalAlignment = -4108 # xlVCenter
            header_rng.HorizontalAlignment = -4108 # xlHAlignCenter
        except Exception as e:
            self.log(f"정렬 설정 중 경고 (무시됨): {e}")

    def apply_symbol_colors(self, ws, max_row, col_idx):
        if col_idx <= 0: return
        symbol_colors = {
            '●': 255,             # Red (FF0000)
            '▲': 26367,           # Orange (FF6600)
            '■': 39168,           # Green (009900)
            '◆': 16711680,        # Blue (0000FF)
            '★': 255,             # Red
            '▼': 8388736,         # Purple (800080)
        }
        for r in range(2, max_row + 1):
            cell = ws.Cells(r, col_idx)
            val = cell.Value
            if val and isinstance(val, str):
                for i, char in enumerate(val):
                    if char in symbol_colors:
                        try:
                            cell.GetCharacters(i+1, 1).Font.Color = symbol_colors[char]
                        except:
                            pass

    def apply_db_formatting(self, db_ws, max_row):
        if not db_ws: return
        c_dict = self.config.get("colors", {})
        try:
            max_c = db_ws.Cells(1, db_ws.Columns.Count).End(-4159).Column
            max_c_let = get_col_letter(max(max_c, 14))

            # 1. Header Row (Row 1): Bold, RGB(226, 239, 218), Center Alignment
            hdr_rng = db_ws.Range(f"A1:{max_c_let}1")
            hdr_rng.Font.Bold = True
            hdr_rng.Interior.Color = hex_to_bgr("#E2EFDA")
            hdr_rng.HorizontalAlignment = -4108 # xlCenter
            hdr_rng.VerticalAlignment = -4108   # xlVCenter

            # 2. Data Rows (Rows 2+): Cols 1..6 Center, Cols 7+ Left Alignment
            if max_row >= 2:
                revision_col = 0
                for column in range(1, max_c + 1):
                    header = str(db_ws.Cells(1, column).Value or "").strip()
                    if header in {"Revision Date", "갱신일"}:
                        revision_col = column
                        break
                if revision_col:
                    from core.db_manager import DBManager
                    revision_range = db_ws.Range(
                        db_ws.Cells(2, revision_col), db_ws.Cells(max_row, revision_col)
                    )
                    revision_range.NumberFormat = "@"
                    for row in range(2, max_row + 1):
                        cell = db_ws.Cells(row, revision_col)
                        normalized = DBManager.normalize_revision_date(cell.Value2)
                        if normalized:
                            cell.Value = normalized

                center_rng = db_ws.Range(f"A2:F{max_row}")
                center_rng.HorizontalAlignment = -4108 # xlCenter
                center_rng.VerticalAlignment = -4108   # xlVCenter

                if max(max_c, 14) >= 7:
                    left_rng = db_ws.Range(f"G2:{max_c_let}{max_row}")
                    left_rng.HorizontalAlignment = -4131 # xlLeft
                    left_rng.VerticalAlignment = -4108   # xlVCenter

                rng = db_ws.Range(f"A2:{max_c_let}{max_row}")

                sf_bg = hex_to_bgr(c_dict.get("search_failed_bg", "#ffff00"))
                sf_fg = hex_to_bgr(c_dict.get("search_failed_font", "#ff0000"))
                mi_bg = hex_to_bgr(c_dict.get("manual_input_bg", "#ffff00"))
                mi_fg = hex_to_bgr(c_dict.get("manual_input_font", "#ff0000"))

                fc_sf = rng.FormatConditions.Add(1, 3, "=\"Search Failed\"")
                fc_sf.Interior.Color = sf_bg
                fc_sf.Font.Color = sf_fg
                fc_sf.Font.Bold = True

                fc_pnf = rng.FormatConditions.Add(1, 3, "=\"Product Not Found\"")
                fc_pnf.Interior.Color = sf_bg
                fc_pnf.Font.Color = sf_fg
                fc_pnf.Font.Bold = True

                fc_mi = rng.FormatConditions.Add(1, 3, "=\"Manual Input Required\"")
                fc_mi.Interior.Color = mi_bg
                fc_mi.Font.Color = mi_fg
                fc_mi.Font.Bold = True

                fc_me = rng.FormatConditions.Add(1, 3, "=\"Manual Entry Required\"")
                fc_me.Interior.Color = mi_bg
                fc_me.Font.Color = mi_fg
                fc_me.Font.Bold = True
        except Exception as e:
            self.log(f"DB 시트 서식 적용 중 경고 (무시됨): {e}")

    def cleanup(self):
        source_wb = self.source_wb
        target_wb = self.target_wb
        self.source_wb = None
        self.target_wb = None
        if target_wb and target_wb != source_wb:
            try:
                target_wb.Close(SaveChanges=False)
            except Exception:
                pass
        if source_wb:
            try:
                source_wb.Close(SaveChanges=False)
            except Exception:
                pass
        if self.excel:
            try:
                self.excel.Quit()
            except Exception:
                pass
            self.excel = None
        self.excel_pid = None
        if hasattr(self, 'sb_context_manager') and self.sb_context_manager:
            try:
                self.sb_context_manager.__exit__(None, None, None)
            except: pass
            self.sb_context_manager = None
            self.sb = None

def refresh_chemical_list_formatting(target_path, config):
    """
    Safely refresh program-managed DB formatting, validation and conditional
    formatting on the latest shared workbook.
    """
    if not target_path or not os.path.exists(target_path):
        return
    import pythoncom
    from core.concurrency_manager import WORKBOOK_OPERATION_LOCK
    if not WORKBOOK_OPERATION_LOCK.acquire(blocking=False):
        raise RuntimeError("다른 동기화 또는 DB 업데이트 작업이 이미 진행 중입니다.")
    pythoncom.CoInitialize()
    engine = SyncEngine(config, callback_progress=lambda msg: print(f"[RefreshFormatting] {msg}"))
    try:
        engine._make_working_copy(target_path)
        return engine._commit_working_copy(
            target_path, changed=True, force_metadata=True
        )
    except Exception as e:
        print(f"[RefreshFormatting Error] {e}")
        raise
    finally:
        engine._discard_working_copy()
        pythoncom.CoUninitialize()
        WORKBOOK_OPERATION_LOCK.release()


def save_db_records_win32com(
    target_path, records_to_save, config, check_stop_fn=None, base_snapshot=None
):
    """
    Saves records_to_save list to 'DB' sheet in target_path using win32com Excel and wb.Save().
    """
    if not target_path or not os.path.exists(target_path) or not records_to_save:
        return
    import win32com.client, pythoncom, datetime
    from core.db_manager import DBManager

    pythoncom.CoInitialize()
    excel = None
    wb = None
    engine = SyncEngine(config or {}, callback_progress=print, check_stop_fn=check_stop_fn)
    try:
        is_auto_sync, max_retries = engine._auto_sync_limits()
        if not check_and_wait_lock(
            target_path, log_fn=print, max_retries=max_retries, retry_delay=60,
            is_auto_sync=is_auto_sync, check_stop_fn=check_stop_fn,
        ):
            raise Exception("ChemicalList.xlsx 파일 사용 대기가 중단되었거나 다음 자동 동기화 주기까지 잠겨 있습니다.")
        work_path = engine._make_working_copy(target_path)
        if base_snapshot is not None:
            # Preserve the state against which crawling decisions were made.
            # The private copy itself is still based on the latest file so all
            # non-overlapping coauthor edits flow through to the commit.
            engine._base_snapshot = base_snapshot

        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.ScreenUpdating = False
        wb = excel.Workbooks.Open(work_path, False, False)

        db_sheet_name = "DB"
        db_ws = None
        for s in wb.Worksheets:
            if s.Name == db_sheet_name:
                db_ws = s
                break
        if not db_ws:
            db_ws = wb.Worksheets.Add(After=wb.Worksheets(wb.Worksheets.Count))
            db_ws.Name = db_sheet_name
            db_cols = DBManager.COLUMNS
            for c_idx, c_name in enumerate(db_cols, 1):
                db_ws.Cells(1, c_idx).Value = c_name

        db_cols_idx = {}
        for c in range(1, 20):
            val = db_ws.Cells(1, c).Value
            if val: db_cols_idx[str(val).strip()] = c
        db_last_col = max(db_cols_idx.values(), default=0)
        for header in DBManager.COLUMNS:
            if header not in db_cols_idx:
                db_last_col += 1
                db_ws.Cells(1, db_last_col).Value = header
                db_cols_idx[header] = db_last_col
        db_last_row = meaningful_data_last_row(db_ws, db_cols_idx, "db")

        db_last_row = db_ws.Cells(db_ws.Rows.Count, 1).End(-4162).Row
        k_db_col = db_cols_idx.get("Key", 1)
        m_db_col = db_cols_idx.get("Manufacturer", 2)
        c_db_col = db_cols_idx.get("Catalog No.", 3)
        rev_db_col = db_cols_idx.get("Revision Date", 14)

        for row_num in range(2, db_last_row + 1):
            revision_cell = db_ws.Cells(row_num, rev_db_col)
            revision = DBManager.normalize_revision_date(revision_cell.Value2)
            revision_cell.NumberFormat = "@"
            if revision:
                revision_cell.Value = revision

        for record in records_to_save:
            norm_dict = {}
            for k, v in record.items():
                norm_k = DBManager.COLUMN_MAP.get(k, k)
                norm_dict[norm_k] = v
            norm_dict["Revision Date"] = datetime.datetime.now().strftime("%Y-%m-%d")
            if "Manufacturer" in norm_dict:
                norm_dict["Manufacturer"] = DBManager.normalize_manufacturer(engine.as_excel_text(norm_dict["Manufacturer"]))
            if "Catalog No." in norm_dict:
                norm_dict["Catalog No."] = engine.as_excel_text(norm_dict["Catalog No."])

            m_val = engine.as_excel_text(norm_dict.get("Manufacturer", "")).lower()
            c_val = engine.as_excel_text(norm_dict.get("Catalog No.", ""))

            target_db_row = 0
            for r in range(2, db_last_row + 1):
                r_man = DBManager.normalize_manufacturer(engine.as_excel_text(db_ws.Cells(r, m_db_col).Value)).lower()
                r_num = engine.as_excel_text(db_ws.Cells(r, c_db_col).Value)
                if r_man == m_val and r_num == c_val:
                    target_db_row = r
                    break
            
            if target_db_row == 0:
                db_last_row += 1
                target_db_row = db_last_row

            db_ws.Cells(target_db_row, m_db_col).NumberFormat = "@"
            db_ws.Cells(target_db_row, c_db_col).NumberFormat = "@"
            db_ws.Cells(target_db_row, rev_db_col).NumberFormat = "@"
            for k, v in norm_dict.items():
                if k == "Key": continue
                c_idx = db_cols_idx.get(k)
                if c_idx:
                    if k in engine.TEXT_FIELDS:
                        v = engine.as_excel_text(v)
                    db_ws.Cells(target_db_row, c_idx).Value = v if v is not None else ""
            db_ws.Cells(target_db_row, k_db_col).Formula = (
                f'={get_col_letter(m_db_col)}{target_db_row}&"|"&'
                f'{get_col_letter(c_db_col)}{target_db_row}'
            )

        if db_last_row >= 2:
            engine.apply_db_formatting(db_ws, db_last_row)

        tgt_ws = None
        for s in wb.Worksheets:
            if s.Name not in ["DB", "가이드", "가이드(Guide)", "Guide", "index", "Old Chemical List"]:
                tgt_ws = s
                break
        if tgt_ws:
            tgt_headers = config.get("target_headers", [])
            tgt_col_map = {}
            tgt_last_col = tgt_ws.Cells(1, tgt_ws.Columns.Count).End(-4159).Column
            for c in range(1, max(tgt_last_col + 1, 30)):
                val = tgt_ws.Cells(1, c).Value
                if val: tgt_col_map[str(val).strip()] = c
            for idx, name in enumerate(tgt_headers):
                if name not in tgt_col_map: tgt_col_map[name] = idx + 1
            max_r = max(2, tgt_ws.Cells(tgt_ws.Rows.Count, 1).End(-4162).Row)
            engine.apply_formatting_and_validation(tgt_ws, tgt_col_map, max_r)

        excel.ScreenUpdating = True
        wb.Save()
        wb.Close(False)
        wb = None
        excel.Quit()
        excel = None
        engine._commit_working_copy(target_path, changed=True)
    except Exception as e:
        print(f"[save_db_records_win32com Error] {e}")
        raise e
    finally:
        if wb:
            try: wb.Close(False)
            except Exception: pass
        if excel:
            try: excel.Quit()
            except Exception: pass
        engine._discard_working_copy()
        try:
            pythoncom.CoUninitialize()
        except Exception: pass

