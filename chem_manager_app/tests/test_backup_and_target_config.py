import datetime as dt
import os
import tempfile
import unittest

from openpyxl import Workbook

from core.backup_manager import create_backup, get_backup_dir, purge_expired_backups, subtract_calendar_months
from core.config_manager import resolve_target_file, validate_chemical_list_file


class BackupRetentionTests(unittest.TestCase):
    def test_calendar_month_cutoff_handles_end_of_month(self):
        value = dt.datetime(2026, 5, 31, 12, 0)
        self.assertEqual(subtract_calendar_months(value, 3), dt.datetime(2026, 2, 28, 12, 0))

    def test_only_matching_backups_at_least_three_months_old_are_deleted(self):
        with tempfile.TemporaryDirectory() as root:
            backup = os.path.join(root, "backup")
            os.mkdir(backup)
            old = os.path.join(backup, "ChemicalList_backup_20260101_000000_000001.xlsx")
            exact = os.path.join(backup, "ChemicalList_backup_20260514_120000_000001.xlsx")
            recent = os.path.join(backup, "ChemicalList_backup_20260801_000000_000001.xlsx")
            unrelated = os.path.join(backup, "other.xlsx")
            for path in (old, exact, recent, unrelated):
                open(path, "wb").close()
            os.utime(old, (dt.datetime(2026, 1, 1).timestamp(),) * 2)
            os.utime(exact, (dt.datetime(2026, 5, 14, 12, 0).timestamp(),) * 2)
            os.utime(recent, (dt.datetime(2026, 8, 1).timestamp(),) * 2)
            os.utime(unrelated, (dt.datetime(2025, 1, 1).timestamp(),) * 2)

            deleted, failures = purge_expired_backups(
                now=dt.datetime(2026, 8, 14, 12, 0), app_root=root
            )

            self.assertEqual(failures, [])
            self.assertEqual({os.path.basename(path) for path in deleted}, {os.path.basename(old), os.path.basename(exact)})
            self.assertFalse(os.path.exists(old))
            self.assertFalse(os.path.exists(exact))
            self.assertTrue(os.path.exists(recent))
            self.assertTrue(os.path.exists(unrelated))

    def test_backup_is_created_under_program_root_not_chemical_list_folder(self):
        with tempfile.TemporaryDirectory() as root:
            program_root = os.path.join(root, "program")
            chemical_dir = os.path.join(root, "onedrive", "shared")
            os.makedirs(program_root)
            os.makedirs(chemical_dir)
            source = os.path.join(chemical_dir, "ChemicalList.xlsx")
            with open(source, "wb") as stream:
                stream.write(b"workbook")
            old_source_time = dt.datetime(2025, 1, 1).timestamp()
            os.utime(source, (old_source_time, old_source_time))
            created_at = dt.datetime(2026, 8, 15, 9, 30)

            backup_path = create_backup(source, now=created_at, app_root=program_root)

            self.assertEqual(os.path.dirname(backup_path), str(get_backup_dir(program_root)))
            self.assertFalse(os.path.exists(os.path.join(chemical_dir, "backup")))
            self.assertEqual(open(backup_path, "rb").read(), b"workbook")
            self.assertEqual(dt.datetime.fromtimestamp(os.path.getmtime(backup_path)), created_at)


class TargetFileTests(unittest.TestCase):
    def test_explicit_target_wins_and_legacy_fallback_remains(self):
        explicit = resolve_target_file({"source_file": r"C:\orders\book.xlsx", "target_file": r"D:\lists\chosen.xlsx"})
        fallback = resolve_target_file({"source_file": r"C:\orders\book.xlsx", "target_file": ""})
        self.assertTrue(explicit.lower().endswith(r"d:\lists\chosen.xlsx"))
        self.assertTrue(fallback.lower().endswith(r"c:\orders\chemicallist.xlsx"))

    def test_target_requires_chemical_list_and_db_sheets(self):
        with tempfile.TemporaryDirectory() as root:
            valid_path = os.path.join(root, "ChemicalList.xlsx")
            invalid_path = os.path.join(root, "other.xlsx")

            valid = Workbook()
            valid.active.title = "ChemicalList"
            valid.create_sheet("DB")
            valid.save(valid_path)
            invalid = Workbook()
            invalid.save(invalid_path)

            self.assertEqual(validate_chemical_list_file(valid_path), (True, ""))
            ok, message = validate_chemical_list_file(invalid_path)
            self.assertFalse(ok)
            self.assertIn("필수 시트", message)


if __name__ == "__main__":
    unittest.main()
