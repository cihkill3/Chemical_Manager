import unittest
import os
import sys
import tempfile

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation

from core.concurrency_manager import (
    SheetSnapshot,
    build_three_way_plan,
    format_validation_max_row,
    workbook_range_needs_extension,
    workbook_main_sheet_needs_activation,
    legacy_order_key,
)
from core.excel_manager import meaningful_data_last_row


def snapshot(kind, rows):
    return {
        kind: SheetSnapshot(
            name="DB" if kind == "db" else "ChemicalList",
            key_kind=kind,
            headers=list(next(iter(rows.values())).keys()) if rows else [],
            rows=rows,
        )
    }


class ThreeWayMergeTests(unittest.TestCase):
    def test_detects_when_workbook_opens_on_db_instead_of_chemical_list(self):
        with tempfile.TemporaryDirectory() as folder:
            path = os.path.join(folder, "active-sheet.xlsx")
            workbook = Workbook()
            chemical = workbook.active
            chemical.title = "ChemicalList"
            db = workbook.create_sheet("DB")
            workbook.active = workbook.sheetnames.index("DB")
            workbook.save(path)
            workbook.close()

            self.assertTrue(workbook_main_sheet_needs_activation(path))

            workbook = Workbook()
            workbook.active.title = "ChemicalList"
            workbook.create_sheet("DB")
            workbook.save(path)
            workbook.close()
            self.assertFalse(workbook_main_sheet_needs_activation(path))

    def test_legacy_rows_without_order_numbers_are_merged_by_fallback_key(self):
        row = {
            "Order No.": "", "Manufacturer": "Aldrich", "Catalog No.": "T1892",
            "Lot No.": "L1", "Product Name": "Legacy reagent", "Remarks": "", "Room": "",
        }
        key = legacy_order_key(row)
        base = snapshot("order", {key: dict(row)})
        ours_row = dict(row, **{"Remarks": "program"})
        theirs_row = dict(row, **{"Room": "501"})
        ours = snapshot("order", {key: ours_row})
        theirs = snapshot("order", {key: theirs_row})

        patches, conflicts = build_three_way_plan(base, ours, theirs)

        self.assertEqual([], conflicts)
        self.assertEqual([("Remarks", "program")], [(p.column, p.new_value) for p in patches])

    def test_sparse_order_number_does_not_hide_existing_data_rows(self):
        class EndCell:
            def __init__(self, row):
                self.Row = row

        class Cell:
            def __init__(self, row):
                self._row = row

            def End(self, _direction):
                return EndCell(self._row)

        class Cells:
            def __init__(self, last_rows):
                self.last_rows = last_rows

            def __call__(self, _row, column):
                return Cell(self.last_rows.get(column, 1))

        class Rows:
            Count = 1048576

        class Worksheet:
            pass

        worksheet = Worksheet()
        worksheet.Rows = Rows()
        worksheet.Cells = Cells({1: 1, 4: 60, 5: 60, 8: 60, 26: 60})

        headers = {
            "Order No.": 1, "Product Name": 4, "Manufacturer": 5,
            "Catalog No.": 8, "Original Product Name": 26,
        }

        self.assertEqual(60, meaningful_data_last_row(worksheet, headers, "order"))

    def test_non_overlapping_coauthor_change_is_preserved(self):
        base = snapshot("order", {"26-1": {"Order No.": "26-1", "Room": "", "Remarks": ""}})
        ours = snapshot("order", {"26-1": {"Order No.": "26-1", "Room": "R1", "Remarks": ""}})
        theirs = snapshot("order", {"26-1": {"Order No.": "26-1", "Room": "", "Remarks": "사용자 입력"}})

        patches, conflicts = build_three_way_plan(base, ours, theirs)

        self.assertEqual([], conflicts)
        self.assertEqual([("Room", "R1")], [(p.column, p.new_value) for p in patches])

    def test_same_cell_change_conflicts(self):
        base = snapshot("order", {"26-1": {"Order No.": "26-1", "Room": ""}})
        ours = snapshot("order", {"26-1": {"Order No.": "26-1", "Room": "R1"}})
        theirs = snapshot("order", {"26-1": {"Order No.": "26-1", "Room": "R2"}})

        patches, conflicts = build_three_way_plan(base, ours, theirs)

        self.assertEqual([], patches)
        self.assertEqual(1, len(conflicts))

    def test_identical_concurrent_result_needs_no_patch(self):
        base = snapshot("db", {"aldrich|1": {"Manufacturer": "Aldrich", "Catalog No.": "1", "CAS No.": ""}})
        ours = snapshot("db", {"aldrich|1": {"Manufacturer": "Aldrich", "Catalog No.": "1", "CAS No.": "50-00-0"}})
        theirs = snapshot("db", {"aldrich|1": {"Manufacturer": "Aldrich", "Catalog No.": "1", "CAS No.": "50-00-0"}})

        patches, conflicts = build_three_way_plan(base, ours, theirs)

        self.assertEqual([], conflicts)
        self.assertEqual([], patches)

    def test_deleted_row_conflicts(self):
        base = snapshot("order", {"26-1": {"Order No.": "26-1", "Room": ""}})
        ours = snapshot("order", {"26-1": {"Order No.": "26-1", "Room": "R1"}})
        theirs = snapshot("order", {})

        _, conflicts = build_three_way_plan(base, ours, theirs)

        self.assertEqual(1, len(conflicts))

    def test_header_structure_change_conflicts(self):
        base = snapshot("order", {"26-1": {"Order No.": "26-1", "Room": ""}})
        ours = snapshot("order", {"26-1": {"Order No.": "26-1", "Room": "R1"}})
        theirs = snapshot("order", {"26-1": {"Order No.": "26-1", "Room": "", "New Column": ""}})

        _, conflicts = build_three_way_plan(base, ours, theirs)

        self.assertEqual(1, len(conflicts))

    def test_program_can_add_db_sheet_when_latest_still_has_none(self):
        ours = snapshot("db", {
            "aldrich|1": {
                "Manufacturer": "Aldrich", "Catalog No.": "1", "CAS No.": "50-00-0"
            }
        })

        patches, conflicts = build_three_way_plan({}, ours, {})

        self.assertEqual([], conflicts)
        self.assertTrue(patches)
        self.assertTrue(all(p.new_row for p in patches))

    def test_validation_range_stops_at_actual_data(self):
        self.assertEqual(3000, format_validation_max_row(1))
        self.assertEqual(3000, format_validation_max_row(2500))
        self.assertEqual(4000, format_validation_max_row(2501))
        self.assertEqual(26500, format_validation_max_row(25001))

    def test_metadata_range_extension_detection(self):
        with tempfile.TemporaryDirectory() as folder:
            path = os.path.join(folder, "range.xlsx")
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "ChemicalList"
            sheet.append(["Order No.", "Room"])
            validation = DataValidation(type="list", formula1='"R1,R2"')
            sheet.add_data_validation(validation)
            validation.add("B2:B20000")
            sheet.conditional_formatting.add(
                "A2:B20000", FormulaRule(formula=['A2=""'])
            )
            workbook.save(path)
            workbook.close()

            self.assertFalse(workbook_range_needs_extension(path, 20000))
            self.assertTrue(workbook_range_needs_extension(path, 25000))

    def test_missing_readable_validation_does_not_force_repeated_save(self):
        """Excel x14 validation can be invisible while managed CF is current."""
        with tempfile.TemporaryDirectory() as folder:
            path = os.path.join(folder, "range-with-x14-like-validation.xlsx")
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "ChemicalList"
            sheet.append(["Order No.", "Room"])
            sheet.append(["26-1", "R1"])
            sheet.conditional_formatting.add(
                "A2:B60", FormulaRule(formula=['A2=""'])
            )
            workbook.save(path)
            workbook.close()

            self.assertFalse(workbook_range_needs_extension(path, 60))
            self.assertTrue(workbook_range_needs_extension(path, 61))


if __name__ == "__main__":
    unittest.main()
