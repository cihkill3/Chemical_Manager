import openpyxl
from openpyxl.styles import PatternFill

# 1. Update order book to have a row with missing CAS/Num
file_path = r"c:\Users\Jeonghun Lee\.gemini\antigravity\scratch\chemical manager\order book.xlsx"
wb = openpyxl.load_workbook(file_path)
ws = wb["2026"]

# Find a row and clear its CAS and Num
for r in range(8, ws.max_row + 1):
    order = ws.cell(row=r, column=8).value
    reag = ws.cell(row=r, column=18).value
    recv = ws.cell(row=r, column=19).value
    if reag in ["O", "o"] and recv in ["O", "o"]:
        # Set 번호 (column 1) to a specific value to trace
        ws.cell(row=r, column=1).value = "99-9999"
        # Clear CAS and Num
        ws.cell(row=r, column=16).value = "" # CAS
        ws.cell(row=r, column=17).value = "" # 품번
        print(f"Set row {r} to miss CAS/Num with 번호 99-9999")
        break
wb.save(file_path)
print("Updated order book.")
