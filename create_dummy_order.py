import openpyxl

file_path = r"c:\Users\Jeonghun Lee\.gemini\antigravity\scratch\chemical manager\order book test.xlsx"
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Sheet1"

# Add dummy headers at row 7
headers = ["번호", "날짜", "주문자", "회사", "품목명", "시약", "수령날짜", "CAS 번호", "품번", "용량", "수량", "보관온도"]
for col, h in enumerate(headers, 1):
    ws.cell(row=7, column=col, value=h)

# Add a row of data at row 8
data = ["12345", "2024-01-01", "홍길동", "Sigma", "NaCl", "O", "2024-01-05", "123-45-6", "S123", "100g", "2", "-20C"]
for col, d in enumerate(data, 1):
    ws.cell(row=8, column=col, value=d)

# Add a row of data at row 9 with NO 수령날짜
data2 = ["12346", "2024-01-02", "홍길동", "Sigma", "KCl", "O", "", "123-45-7", "S124", "100g", "1", "RT"]
for col, d in enumerate(data2, 1):
    ws.cell(row=9, column=col, value=d)

wb.save(file_path)
print("Created dummy order book test.xlsx")
