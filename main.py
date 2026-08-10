import os
import pandas as pd
from db_manager import DBManager
import importlib
from sds_downloader import download_thermofisher_sds, download_aldrich_sds, download_tci_sds
from hazard_parser import parse_hazard
from openpyxl.styles import Font, Alignment

# SeleniumBase 임포트 (Playwright 대체)
try:
    from seleniumbase import Driver
except ImportError:
    print("SeleniumBase가 설치되어 있지 않습니다. pip install seleniumbase undetected-chromedriver 를 실행하세요.")
    Driver = None

SUPPORTED_MANUFACTURERS = {
    "thermo fisher": "thermofisher",
    "thermofisher": "thermofisher",
    "aldrich": "aldrich",
    "sigma aldrich": "aldrich",
    "sigma-aldrich": "aldrich",
    "tci": "tci",
    "abcam": "abcam"
}

def process_orderbook(input_excel_path, output_excel_path, db_path="chemical_db.xlsx"):
    db = DBManager(db_path)
    
    try:
        df_order = pd.read_excel(input_excel_path)
    except Exception as e:
        print(f"Error reading input file: {e}")
        return

    # 결과물에 출력할 컬럼 추가
    for col in db.columns:
        if col not in df_order.columns and col not in ["제조사", "제품번호", "갱신일"]:
            df_order[col] = ""

    if Driver is None:
        return

    # SeleniumBase UC Mode (Undetected Chromedriver) 드라이버 생성
    # 헤드리스 모드로 실행하되 봇 탐지 우회를 위해 uc=True 적용
    print("Initializing browser (UC Mode)...")
    driver = Driver(uc=True, headless=True)
    driver.set_page_load_timeout(30)
    
    scrapers = {}
    
    def get_scraper(manufacturer_key):
        if manufacturer_key not in scrapers:
            try:
                module_name = f"scrapers.{manufacturer_key}"
                module = importlib.import_module(module_name)
                class_name = f"{manufacturer_key.capitalize()}Scraper"
                scraper_class = getattr(module, class_name)
                scrapers[manufacturer_key] = scraper_class(driver)
            except Exception as e:
                print(f"Failed to load scraper for {manufacturer_key}: {e}")
                return None
        return scrapers[manufacturer_key]

    try:
        for index, row in df_order.iterrows():
            manufacturer_raw = str(row.get("제조사", "")).strip()
            product_number = str(row.get("제품번호", "")).strip()

            if not manufacturer_raw or not product_number or manufacturer_raw == "nan":
                continue

            print(f"\nProcessing: {manufacturer_raw} - {product_number}")

            # 1. DB 먼저 확인
            db_record = db.get_product(manufacturer_raw, product_number)
            if db_record:
                print(" -> Found in DB")
                
                # DB 호환성 유지: 과거 '위험분류' 컬럼이 있는 경우 파싱해서 새 컬럼에 채움
                if "위험분류" in db_record and ("신호어" not in db_record or pd.isna(db_record.get("신호어", "")) or db_record.get("신호어", "") == ""):
                    hazard_str = db_record.pop("위험분류")
                    s_word, m_hazard, d_hazard = parse_hazard(hazard_str)
                    db_record["신호어"] = s_word
                    db_record["주요위험"] = m_hazard
                    db_record["상세 위험분류"] = d_hazard
                
                for key in db.columns:
                    if key in df_order.columns and key not in ["제조사", "제품번호"]:
                        df_order.at[index, key] = db_record.get(key, "정보 없음")
                continue

            # 2. 지원하는 제조사인지 확인
            m_lower = manufacturer_raw.lower()
            manufacturer_key = None
            for key, val in SUPPORTED_MANUFACTURERS.items():
                if key in m_lower:
                    manufacturer_key = val
                    break
            
            if not manufacturer_key:
                print(" -> Unsupported manufacturer")
                df_order.at[index, "시약명"] = "직접 입력하세요"
                continue

            # 3. 크롤러 동작
            scraper = get_scraper(manufacturer_key)
            if scraper:
                print(f" -> Scraping via {manufacturer_key}...")
                scraped_data = scraper.scrape(product_number)
                
                if scraped_data:
                    # 제조사명은 오더북에 적힌 원본 유지
                    scraped_data["제조사"] = manufacturer_raw
                    
                    if "위험분류" in scraped_data:
                        hazard_str = scraped_data.pop("위험분류")
                        s_word, m_hazard, d_hazard = parse_hazard(hazard_str)
                        scraped_data["신호어"] = s_word
                        scraped_data["주요위험"] = m_hazard
                        scraped_data["상세 위험분류"] = d_hazard
                        
                    print(f" -> Downloading SDS for {product_number}...")
                    sds_dir = "sds_downloads"
                    sds_path = ""
                    if manufacturer_key == "thermofisher":
                        sds_path = download_thermofisher_sds(product_number, sds_dir)
                    elif manufacturer_key == "aldrich":
                        sds_path = download_aldrich_sds(product_number, "aldrich", sds_dir)
                    elif manufacturer_key == "tci":
                        sds_path = download_tci_sds(product_number, sds_dir)
                        
                    if sds_path:
                        scraped_data["sds"] = sds_path
                    else:
                        scraped_data["sds"] = "정보 없음"
                    
                    db.add_product(scraped_data)
                    for key in db.columns:
                        if key in df_order.columns and key not in ["제조사", "제품번호"]:
                            val = scraped_data.get(key)
                            if not val or str(val).strip().lower() in ["", "n/a", "na", "none", "nan"]:
                                val = "정보 없음"
                            df_order.at[index, key] = val
                else:
                    print(" -> Scrape failed or product not found.")
                    # 검색 실패시 명시된 문구로 처리
                    df_order.at[index, "시약명"] = "제조사 홈페이지에서 검색 실패"
            else:
                print(" -> Scraper module not found")
                
    finally:
        print("\nClosing browser...")
        driver.quit()

    # 결과 저장 시 이모지가 잘 보이도록 포맷팅
    with pd.ExcelWriter(output_excel_path, engine='openpyxl') as writer:
        df_order.to_excel(writer, index=False)
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']
        
        from openpyxl.cell.rich_text import CellRichText, TextBlock
        from openpyxl.cell.text import InlineFont
        import re
        
        wrap_alignment = Alignment(wrap_text=True, vertical='center')
        
        symbol_colors = {
            '●': 'FF0000', # Red
            '▲': 'FF6600', # Orange
            '■': '009900', # Green
            '◆': '0000FF', # Blue
            '★': 'FF0000', # Red
            '▼': '800080', # Purple
        }
        
        # 컬럼 인덱스 찾기
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1))
        signal_col_idx = None
        main_hazard_col_idx = None
        
        for idx, cell in enumerate(header_row):
            if cell.value == '신호어':
                signal_col_idx = idx + 1
            elif cell.value == '주요위험':
                main_hazard_col_idx = idx + 1
                
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.alignment = wrap_alignment
                if cell.column in [signal_col_idx, main_hazard_col_idx] and isinstance(cell.value, str):
                    val = cell.value
                    if val and val != "정보 없음":
                        rich_elements = []
                        parts = re.split(r'([●▲■◆★▼])', val)
                        for part in parts:
                            if not part: continue
                            if part in symbol_colors:
                                rich_elements.append(TextBlock(InlineFont(color=symbol_colors[part]), part))
                            else:
                                rich_elements.append(part)
                        if rich_elements:
                            cell.value = CellRichText(*rich_elements)
                    
        # 컬럼 너비 조정
        for col_idx, column in enumerate(worksheet.columns, 1):
            col_name = header_row[col_idx-1].value
            if col_name == '시약명':
                worksheet.column_dimensions[column[0].column_letter].width = 45
            elif col_name == '상세 위험분류':
                worksheet.column_dimensions[column[0].column_letter].width = 50
            elif col_name in ['주요위험', '민감성']:
                worksheet.column_dimensions[column[0].column_letter].width = 25
            else:
                worksheet.column_dimensions[column[0].column_letter].width = 15

    print(f"Processing complete. Result saved to {output_excel_path}")

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument('--input', type=str, default='sample_order.xlsx')
    parser.add_argument('--output', type=str, default='result.xlsx')
    parser.add_argument('--db', type=str, default='chemical_db.xlsx')
    args = parser.parse_args()
    
    if not os.path.exists(args.input):
        print(f"{args.input}이 없습니다.")
    else:
        process_orderbook(args.input, args.output, args.db)
