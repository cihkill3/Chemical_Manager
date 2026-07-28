import openpyxl

file_path = r"c:\Users\Jeonghun Lee\.gemini\antigravity\scratch\chemical manager\order book.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
ws = wb["2026"]

# 헤더 확인
print("Header at row 7:")
headers = []
for c in range(1, 20):
    headers.append(str(ws.cell(row=7, column=c).value or "").strip())
print(headers)

valid_count = 0
for row_idx in range(8, ws.max_row + 1):
    order_num = str(ws.cell(row=row_idx, column=8).value or "").strip()
    reagent = str(ws.cell(row=row_idx, column=18).value or "").strip()
    received = str(ws.cell(row=row_idx, column=19).value or "").strip()
    
    if (reagent in ["O", "o", "Ｏ"]) and (received in ["O", "o", "Ｏ"]):
        print(f"Row {row_idx} : OrderNum='{order_num}', Reagent='{reagent}', Received='{received}'")
        if order_num:
            valid_count += 1

print(f"Total valid rows WITH order_num: {valid_count}")
