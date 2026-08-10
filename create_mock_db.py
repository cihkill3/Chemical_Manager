import pandas as pd

# 실제 다운로드된 SDS 문서에 기반한 정보
data = [
    {
        '제조사': 'Thermo Fisher', 
        '제품번호': 'L16400', 
        '제품명': '3-(Acryloyloxy)propyltrimethoxysilane, 94%', 
        '신호어': '🔴 위험', 
        '주요위험 (요약)': '☠️급성독성, 🧪부식성', 
        '상세 위험분류 (SDS 원문 내용)': '급성 독성(경구/경피/흡입) - 구분 3\n피부 부식성/자극성 - 구분 1B\n피부 과민성 - 구분 1'
    },
    {
        '제조사': 'TCI', 
        '제품번호': 'C0119', 
        '제품명': "1,1'-Carbonyldiimidazole", 
        '신호어': '🔴 위험', 
        '주요위험 (요약)': '☠️독성, 🐟환경유해', 
        '상세 위험분류 (SDS 원문 내용)': '급성 독성(흡입) - 구분 3\n수생환경 유해성(만성) - 구분 2'
    },
    {
        '제조사': 'Thermo Fisher', 
        '제품번호': 'L09319', 
        '제품명': 'Fluorescein isothiocyanate, isomer 1, 95%', 
        '신호어': '🟡 경고', 
        '주요위험 (요약)': '⚠️자극성', 
        '상세 위험분류 (SDS 원문 내용)': '심한 눈 손상성/자극성 - 구분 2A'
    },
    {
        '제조사': 'TCI', 
        '제품번호': 'T0751', 
        '제품명': 'Trifluoromethanesulfonic Acid', 
        '신호어': '🔴 위험', 
        '주요위험 (요약)': '🧪부식성, 👤건강유해', 
        '상세 위험분류 (SDS 원문 내용)': '피부 부식성 - 구분 1A\n특정표적장기 독성(반복 노출) - 구분 2'
    },
    {
        '제조사': 'Thermo Fisher', 
        '제품번호': '300-13-50UG', 
        '제품명': 'Recombinant Protein (FGF-basic)', 
        '신호어': '🟢 안전', 
        '주요위험 (요약)': '🟢해당없음', 
        '상세 위험분류 (SDS 원문 내용)': 'GHS 기준에 의거하여 유해화학물질로 분류되지 않음'
    }
]

df = pd.DataFrame(data)

# 엑셀 파일 생성 시 이모지가 잘 보이도록 폰트 지정 (Segoe UI Emoji)
with pd.ExcelWriter('example_hazard_db.xlsx', engine='openpyxl') as writer:
    df.to_excel(writer, index=False, sheet_name='Sheet1')
    
    workbook = writer.book
    worksheet = writer.sheets['Sheet1']
    
    from openpyxl.styles import Font, Alignment
    
    # 이모지용 폰트 (Windows 기준 가장 이모지 호환성이 좋은 폰트)
    emoji_font = Font(name='Segoe UI Emoji', size=11)
    # 줄바꿈 및 자동 맞춤을 위한 설정
    wrap_alignment = Alignment(wrap_text=True, vertical='center')
    
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
        for cell in row:
            # 기본 정렬(가운데 맞춤 및 줄바꿈)
            cell.alignment = wrap_alignment
            # 신호어 및 주요위험 컬럼(D열, E열)은 이모지 전용 폰트 적용
            if cell.column_letter in ['D', 'E']:
                cell.font = emoji_font
    
    # 컬럼 너비 조정
    worksheet.column_dimensions['A'].width = 15
    worksheet.column_dimensions['B'].width = 15
    worksheet.column_dimensions['C'].width = 45
    worksheet.column_dimensions['D'].width = 12
    worksheet.column_dimensions['E'].width = 25
    worksheet.column_dimensions['F'].width = 50
